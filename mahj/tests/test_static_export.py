"""Static export of the public spectator page.

Exercises mahj.publish.static_export end-to-end against the seeded tournament:
the right files are produced, modal links are rewritten to resolvable .html
targets, the live WebSocket client is swapped for the poller, and the DB-served
tenant logo is emitted as a file. Reveal masking is inherited from the anonymous
views (round 3 is unpublished in the fixture) and asserted here too.
"""
import json

import pytest

from mahj.publish.static_export import export_public, _team_slug
from mahj.tests.conftest import grant


@pytest.fixture
def teamed_tournament(tournament):
    """Give players two teams and the tenant a logo, to cover those export branches."""
    players = tournament['players']
    for i, p in enumerate(players):
        p.team = 'Alpha' if i % 2 == 0 else 'Beta'
        p.save()
    v = tournament['settings']
    v.has_teams = True
    v.logo = b'\x89PNG\r\n\x1a\n fake-png-bytes'
    v.logo_etag = 'deadbeef'
    v.save()
    return tournament


def _export(tmp_path, subdomain='test'):
    out = tmp_path / subdomain
    export_public(subdomain, out, copy_static=False)
    return out


class _FakeSFTP:
    """Records what an upload did instead of talking to a server: files written,
    directories created, files removed, and what listdir reports."""

    def __init__(self):
        self.written = {}
        self.made_dirs = set()
        self.removed = []
        self.listing = []

    # -- the paramiko SFTPClient surface upload_dump/upload_dir use ------------
    def stat(self, path):
        if path not in self.made_dirs:
            raise IOError(f'no such dir {path}')

    def mkdir(self, path):
        self.made_dirs.add(path)

    def putfo(self, fileobj, remote):
        self.written[remote] = fileobj.read()
        # A real server lists what was just written, which is what the pruning
        # that follows an upload sees.
        self.listing.append(remote.rsplit('/', 1)[-1])

    def put(self, local, remote):
        self.written[remote] = open(local, 'rb').read()

    def listdir(self, path):
        return list(self.listing)

    def remove(self, path):
        self.removed.append(path)

    def close(self):
        pass


@pytest.fixture
def fake_sftp(monkeypatch):
    """Stand in for a live SFTP connection (see test_sftp_host_key for the
    host-key half of this)."""
    sftp = _FakeSFTP()

    class _Client:
        def open_sftp(self):
            return sftp

        def close(self):
            pass

    monkeypatch.setattr('mahj.publish.sftp_upload._connect', lambda cfg: _Client())
    return sftp


class TestFilesProduced:
    def test_index_and_version(self, tmp_path, tournament):
        out = _export(tmp_path)
        assert (out / 'index.html').exists()
        version = json.loads((out / 'version.json').read_text())
        assert 'version' in version

    def test_player_and_score_modals(self, tmp_path, tournament):
        out = _export(tmp_path)
        pid = tournament['players'][0].id
        assert (out / f'details_player_{pid}.html').exists()
        # Round 1 table 1 exists in the fixture.
        assert (out / 'detailed_scores_1_1.html').exists()

    def test_team_modal_uses_slug(self, tmp_path, teamed_tournament):
        out = _export(tmp_path)
        assert (out / f'details_team_{_team_slug("Alpha")}.html').exists()
        assert (out / f'details_team_{_team_slug("Beta")}.html').exists()

    def test_stats_xlsx_produced_and_linked(self, tmp_path, tournament):
        import io
        from openpyxl import load_workbook
        out = _export(tmp_path)
        f = out / 'stats.xlsx'
        assert f.exists()
        wb = load_workbook(io.BytesIO(f.read_bytes()))
        headers = [c.value for c in wb['Players'][1]]
        assert 'Rank' in headers and 'Player' in headers
        assert wb['Players'].max_row >= 2   # header + at least one player
        # The stats-tab button links to it; the href stays relative (not .html-suffixed
        # like the modal links) so it resolves to the exported file in the flat dir.
        assert 'href="stats.xlsx"' in (out / 'index.html').read_text()


class TestRewrites:
    def test_modal_links_get_html_suffix(self, tmp_path, tournament):
        out = _export(tmp_path)
        html = (out / 'index.html').read_text()
        pid = tournament['players'][0].id
        # A literal player link now points at the exported file.
        assert f'details_player_{pid}.html' in html
        # No bare (extension-less) modal href should survive.
        assert f'href="details_player_{pid}"' not in html

    def test_team_link_rewritten_to_slug(self, tmp_path, teamed_tournament):
        out = _export(tmp_path)
        html = (out / 'index.html').read_text()
        assert f'details_team_{_team_slug("Alpha")}.html' in html
        # The raw team name is gone from the href.
        assert 'href="details_team_Alpha"' not in html

    def test_static_urls_are_relative(self, tmp_path, tournament):
        # So the site can be hosted in a subfolder, not just at the domain root.
        html = (_export(tmp_path) / 'index.html').read_text()
        assert '"/static/' not in html          # no root-absolute static URLs
        assert 'src="static/js/static_poll.js"' in html

    def test_websocket_client_swapped_for_poller(self, tmp_path, tournament):
        out = _export(tmp_path)
        html = (out / 'index.html').read_text()
        assert 'display_socket' not in html
        assert 'static_poll.js' in html

    def test_logo_blob_written_and_referenced(self, tmp_path, teamed_tournament):
        out = _export(tmp_path)
        assert (out / 'logo.png').exists()
        html = (out / 'index.html').read_text()
        assert 'logo.png' in html
        assert '/logo?v=' not in html

    def test_no_logo_keeps_static_fallback(self, tmp_path, tournament):
        # Fixture tenant has no uploaded logo → the static mcr_logo stays, no file.
        out = _export(tmp_path)
        assert not (out / 'logo.png').exists()


class TestAuthMenu:
    def test_export_drops_login_menu(self, tmp_path, tournament):
        # No auth on a static host — the overflow menu and its login link are gone.
        html = (_export(tmp_path) / 'index.html').read_text()
        assert 'accounts/login' not in html
        assert 'open = !open' not in html   # the ⋮ overflow-menu toggle

    def test_export_drops_live_indicator(self, tmp_path, tournament):
        # No live socket on a static snapshot — the Live/Offline pill is gone.
        html = (_export(tmp_path) / 'index.html').read_text()
        assert 'x-show="wsConnected"' not in html

    def test_live_anon_view_keeps_login(self, tournament):
        # Regression guard: the normal (served) anon page still offers Login.
        from django.test import Client
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        html = c.get('/').content.decode()
        assert 'accounts/login' in html


class TestPublishProgress:
    def test_status_requires_staff(self, tournament):
        from django.test import Client
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        resp = c.get('/publish_status')
        assert resp.status_code in (301, 302)   # anonymous → login

    def test_status_idle_by_default(self, tournament):
        from django.contrib.auth.models import User
        from django.test import Client
        u = User.objects.create_user('boss', password='pw')
        grant(u, tournament['tenant'], admin=True)
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        c.force_login(u)
        assert c.get('/publish_status').json()['phase'] == 'idle'

    def test_progress_round_trip(self, settings):
        # The real cache (not the test DummyCache) is what the daemon thread and
        # the poll endpoint share; verify set/get under a working backend.
        settings.CACHES = {'default': {
            'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}}
        from django.core.cache import caches
        caches['default'].clear()
        from mahj.publish import trigger
        assert trigger.get_progress('t') is None
        trigger.set_progress('t', 'upload', pct=42, message='Uploading… 3/7')
        p = trigger.get_progress('t')
        assert p['phase'] == 'upload' and p['pct'] == 42


class TestRevealMasking:
    def test_unpublished_round_hidden(self, tmp_path, tournament):
        # Round 3 is unpublished in the fixture; the anonymous export must not
        # reveal its scores in the per-table modal.
        out = _export(tmp_path)
        html = (out / 'detailed_scores_3_1.html').read_text()
        # The "not revealed" placeholder stands in for the withheld grid.
        assert 'id="by_1"' not in html


@pytest.mark.django_db
class TestTriggerGating:
    """fire_static_export must stay a no-op unless the tenant has an enabled
    publish target."""

    def _target(self, **kw):
        from mahj.models import Tenant, PublishTarget
        t, _ = Tenant.objects.get_or_create(subdomain='live', defaults={'name': 'Live'})
        return PublishTarget.objects.create(tenant=t, **kw)

    def test_no_target_is_noop(self):
        from mahj.publish import trigger
        assert trigger._should_publish('live') is False

    def test_enabled_target_publishes(self):
        from mahj.publish import trigger
        self._target(enabled=True, host='host.example')
        assert trigger._should_publish('live') is True

    def test_disabled_target_is_noop(self):
        from mahj.publish import trigger
        self._target(enabled=False, host='host.example')
        assert trigger._should_publish('live') is False


class TestSecrets:
    def test_round_trip(self):
        from mahj.publish import secrets
        token = secrets.encrypt('hunter2')
        assert token is not None and bytes(token) != b'hunter2'
        assert secrets.decrypt(token) == 'hunter2'

    def test_empty_is_none_or_blank(self):
        from mahj.publish import secrets
        assert secrets.encrypt('') is None
        assert secrets.decrypt(None) == ''


@pytest.mark.django_db
class TestPublishTargetResolution:
    """resolve_config returns a tenant's enabled PublishTarget, or None."""

    def _target(self, **kw):
        from mahj.models import Tenant, PublishTarget
        t, _ = Tenant.objects.get_or_create(subdomain='acme', defaults={'name': 'Acme'})
        return PublishTarget.objects.create(tenant=t, **kw)

    def test_enabled_target_resolves(self):
        from mahj.publish.sftp_upload import resolve_config
        self._target(enabled=True, host='db.example', path='/srv', username='u')
        cfg = resolve_config('acme')
        assert cfg.host == 'db.example' and cfg.path == '/srv' and cfg.username == 'u'

    def test_disabled_target_is_none(self):
        from mahj.publish.sftp_upload import resolve_config
        self._target(enabled=False, host='db.example')
        assert resolve_config('acme') is None

    def test_enabled_but_no_host_is_none(self):
        from mahj.publish.sftp_upload import resolve_config
        self._target(enabled=True, host='')
        assert resolve_config('acme') is None

    def test_no_target_is_none(self):
        from mahj.publish.sftp_upload import resolve_config
        assert resolve_config('acme') is None

    def test_secret_decrypted_into_config(self):
        from mahj.publish import secrets
        from mahj.publish.sftp_upload import resolve_config
        self._target(enabled=True, host='db.example',
                     password_enc=secrets.encrypt('s3cret'))
        assert resolve_config('acme').password == 's3cret'


class TestPublishWebEndpoint:
    def test_requires_staff(self, tournament):
        from django.test import Client
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        resp = c.post('/publish_web')
        # Anonymous → redirected to login, not a 200/OK publish.
        assert resp.status_code in (301, 302)

    def test_reports_when_unconfigured(self, tournament):
        # Tenant 'test' has no publish target → 400.
        from django.contrib.auth.models import User
        from django.test import Client
        u = User.objects.create_user('boss', password='pw')
        grant(u, tournament['tenant'], admin=True)
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        c.force_login(u)
        resp = c.post('/publish_web')
        assert resp.status_code == 400
        assert 'not configured' in resp.json()['error'].lower()


class TestPublishTargetEndpoint:
    def _staff_client(self):
        from django.contrib.auth.models import User
        from django.test import Client
        from mahj.models import Tenant
        u = User.objects.create_user('boss', password='pw')
        grant(u, Tenant.objects.get(subdomain='test'), admin=True)
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        c.force_login(u)
        return c

    def test_save_requires_staff(self, tournament):
        from django.test import Client
        c = Client()
        c.defaults['HTTP_HOST'] = 'test.example.com'
        resp = c.post('/publish_target_save', {'host': 'x'})
        assert resp.status_code in (301, 302)  # anonymous → login

    def test_save_encrypts_password(self, tournament):
        from mahj.models import PublishTarget
        from mahj.publish import secrets
        c = self._staff_client()
        resp = c.post('/publish_target_save', {
            'enabled': 'true', 'host': 'web.example', 'port': '22',
            'username': 'u', 'path': '/srv', 'password': 'topsecret',
        })
        assert resp.json()['status'] == 'ok'
        t = PublishTarget.objects.get(tenant=tournament['tenant'])
        assert t.enabled and t.host == 'web.example' and t.path == '/srv'
        # Stored ciphertext, not the plaintext, and decrypts back.
        assert bytes(t.password_enc) != b'topsecret'
        assert secrets.decrypt(t.password_enc) == 'topsecret'

    def test_blank_password_keeps_existing(self, tournament):
        from mahj.models import PublishTarget
        from mahj.publish import secrets
        c = self._staff_client()
        c.post('/publish_target_save', {'host': 'web.example', 'password': 'keepme'})
        c.post('/publish_target_save', {'host': 'web2.example'})  # no password sent
        t = PublishTarget.objects.get(tenant=tournament['tenant'])
        assert t.host == 'web2.example'
        assert secrets.decrypt(t.password_enc) == 'keepme'

    def test_save_persists_public_url_to_settings(self, tournament):
        from mahj.models import TournamentSettings
        c = self._staff_client()
        c.post('/publish_target_save',
               {'host': 'web.example', 'public_url': 'https://scores.example.org'})
        v = TournamentSettings.objects.get(tenant=tournament['tenant'])
        assert v.public_url == 'https://scores.example.org'

    def test_test_connection_needs_host(self, tournament):
        # Test connection uses the posted form values, so an empty host is a 400
        # before any network attempt.
        c = self._staff_client()
        resp = c.post('/publish_target_test', {'host': ''})
        assert resp.status_code == 400

    def test_clear_password_wipes_it(self, tournament):
        from mahj.models import PublishTarget
        c = self._staff_client()
        c.post('/publish_target_save', {'host': 'web.example', 'password': 'keepme'})
        c.post('/publish_target_save', {'host': 'web.example', 'clear_password': '1'})
        t = PublishTarget.objects.get(tenant=tournament['tenant'])
        assert t.password_enc is None

    def test_bad_port_is_rejected(self, tournament):
        c = self._staff_client()
        resp = c.post('/publish_target_save', {'host': 'web.example', 'port': 'nope'})
        assert resp.status_code == 400

    def test_page_never_renders_secret(self, tournament):
        c = self._staff_client()
        c.post('/publish_target_save', {'host': 'web.example', 'password': 'topsecret'})
        resp = c.get('/admin?page=publish_target')
        assert resp.status_code == 200
        assert b'topsecret' not in resp.content
        assert b'configured' in resp.content


class TestDumpUpload:
    """Every publish uploads a tournament dump (see tenant_dump), so a published
    state can always be restored — and never into the served tree."""

    def _target(self, tenant, **kw):
        from mahj.models import PublishTarget
        fields = dict(enabled=True, host='web.example', username='u', path='/srv/site')
        fields.update(kw)
        return PublishTarget.objects.create(tenant=tenant, **fields)

    def test_the_default_dump_dir_is_login_relative(self, tournament):
        """A dump under the web root is fetchable by anyone guessing the name, and
        it holds every score including the withheld final round. The login
        directory is not served."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        assert sftp_upload.resolve_config('test').dump_dir() == 'mahj-backups'

    @pytest.mark.parametrize('path', [
        'public_html/2026',      # one tournament a year: the sibling would be
                                 # public_html/mahj-backups, inside the docroot
        '/home/u/public_html/spring',
        'public_html',
        '/srv/site',
        '',
    ])
    def test_no_default_dump_dir_is_derived_from_the_site_path(self, tournament, path):
        """Whatever the target path, the default never lands under it — the config
        never says where the docroot begins, so deriving a directory from `path`
        cannot tell a docroot from a subfolder of one."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'], path=path)
        dump_dir = sftp_upload.resolve_config('test').dump_dir()
        assert dump_dir == 'mahj-backups'
        site = path.rstrip('/')
        assert not (site and dump_dir.startswith(site + '/'))

    def test_backup_path_overrides_it(self, tournament):
        """An operator who wants a specific directory — including one inside the
        site, if they mean it — names it."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'], backup_path='/var/backups/mahj/')
        assert sftp_upload.resolve_config('test').dump_dir() == '/var/backups/mahj'

    def test_upload_dump_sends_the_file_and_prunes_old_ones(self, tournament, fake_sftp):
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        # More existing dumps than we keep, plus another tenant's — which must
        # survive, since tenants may share one backup directory.
        existing = [f'mahj_test_202607{i + 1:02d}T100000Z.json.gz' for i in range(25)]
        # Another tenant's dumps, and one whose subdomain merely *starts* with
        # this one's name — neither may be reaped.
        fake_sftp.listing = existing + ['mahj_other_20260101T100000Z.json.gz',
                                        'mahj_test_2026_20260101T100000Z.json.gz']

        sftp_upload.upload_dump('test', b'payload', 'mahj_test_20260820T120000Z.json.gz')

        assert fake_sftp.written['mahj-backups/mahj_test_20260820T120000Z.json.gz'] == b'payload'
        assert 'mahj-backups' in fake_sftp.made_dirs
        # 25 old + 1 new = 26, keep 20 → the 6 oldest of this tenant's go.
        assert len(fake_sftp.removed) == 6
        assert all('mahj_test_2026_' not in p for p in fake_sftp.removed)
        assert all('mahj_other_' not in p for p in fake_sftp.removed)

    def test_upload_dump_without_a_target_is_a_noop(self, tournament, fake_sftp):
        from mahj.publish import sftp_upload
        sftp_upload.upload_dump('test', b'payload', 'dump.json.gz')
        assert fake_sftp.written == {}

    def test_publish_uploads_a_dump_and_reports_done(self, tournament, monkeypatch, settings):
        """The whole background job: render, upload, then back up."""
        settings.CACHES = {'default': {
            'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}}
        from django.core.cache import caches
        caches['default'].clear()
        from mahj.publish import trigger
        sent = {}
        monkeypatch.setattr(trigger, 'set_progress', lambda *a, **kw: None, raising=False)
        monkeypatch.setattr('mahj.publish.static_export.export_public',
                            lambda *a, **kw: None)
        monkeypatch.setattr('mahj.publish.sftp_upload.upload_dir', lambda *a, **kw: None)
        monkeypatch.setattr('mahj.publish.sftp_upload.upload_dump',
                            lambda sub, data, name: sent.update(sub=sub, data=data, name=name))
        trigger._run('test')
        assert sent['sub'] == 'test' and sent['name'].startswith('mahj_test_')
        # A real dump of the fixture tenant, not an empty one.
        import gzip
        import json as _json
        payload = _json.loads(gzip.decompress(sent['data']))
        assert len(payload['models']['Player']) == 16

    def test_a_failed_backup_still_reports_the_site_as_published(self, tournament, monkeypatch,
                                                                 settings):
        """The site is already live by then, so a backup problem is a note on a
        done publish, never a failed publish."""
        settings.CACHES = {'default': {
            'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}}
        from django.core.cache import caches
        caches['default'].clear()
        from mahj.publish import trigger

        def boom(*a, **kw):
            raise RuntimeError('remote disk full')

        monkeypatch.setattr('mahj.publish.static_export.export_public', lambda *a, **kw: None)
        monkeypatch.setattr('mahj.publish.sftp_upload.upload_dir', lambda *a, **kw: None)
        monkeypatch.setattr('mahj.publish.sftp_upload.upload_dump', boom)
        trigger._run('test')
        progress = trigger.get_progress('test')
        assert progress['phase'] == 'done'
        assert 'remote disk full' in progress['error']


class TestIncrementalUpload:
    """`upload_dir` skips files whose size+mtime match the last upload, recorded in
    a local manifest. Publishing runs on every score change during an event, so the
    skip is what keeps a publish to a couple of seconds instead of re-sending every
    flag SVG — and version.json must always go, since it is the signal that tells a
    spectator's browser to refresh.
    """

    def _target(self, tenant, **kw):
        from mahj.models import PublishTarget
        fields = dict(enabled=True, host='web.example', username='u', path='/srv/site')
        fields.update(kw)
        return PublishTarget.objects.create(tenant=tenant, **fields)

    def _site(self, tmp_path):
        (tmp_path / 'index.html').write_text('<html>')
        (tmp_path / 'assets').mkdir()
        (tmp_path / 'assets' / 'flag.svg').write_text('<svg>')
        (tmp_path / 'version.json').write_text('{"v": 1}')
        return tmp_path

    def test_the_first_upload_sends_everything(self, tournament, fake_sftp, tmp_path):
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        sftp_upload.upload_dir(self._site(tmp_path), 'test')
        assert set(fake_sftp.written) == {
            '/srv/site/index.html', '/srv/site/assets/flag.svg',
            '/srv/site/version.json'}

    def test_version_json_lands_last(self, tournament, fake_sftp, tmp_path):
        """Ordered on purpose: a client that saw a new version before the files it
        points at would fetch a half-published site."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        sftp_upload.upload_dir(self._site(tmp_path), 'test')
        assert list(fake_sftp.written)[-1] == '/srv/site/version.json'

    def test_a_second_upload_skips_unchanged_files(self, tournament, fake_sftp,
                                                   tmp_path):
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        site = self._site(tmp_path)
        sftp_upload.upload_dir(site, 'test')
        fake_sftp.written.clear()
        sftp_upload.upload_dir(site, 'test')
        # Only the refresh signal, not the flag or the page.
        assert set(fake_sftp.written) == {'/srv/site/version.json'}

    def test_a_changed_file_is_sent_again(self, tournament, fake_sftp, tmp_path):
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        site = self._site(tmp_path)
        sftp_upload.upload_dir(site, 'test')
        fake_sftp.written.clear()
        (site / 'index.html').write_text('<html>changed and longer')
        sftp_upload.upload_dir(site, 'test')
        assert '/srv/site/index.html' in fake_sftp.written

    def test_full_ignores_the_manifest(self, tournament, fake_sftp, tmp_path):
        """The operator's escape hatch when a remote file was changed or lost
        behind our back — the manifest would otherwise say it is still fine."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        site = self._site(tmp_path)
        sftp_upload.upload_dir(site, 'test')
        fake_sftp.written.clear()
        sftp_upload.upload_dir(site, 'test', full=True)
        assert len(fake_sftp.written) == 3

    def test_a_corrupt_manifest_falls_back_to_a_full_upload(self, tournament,
                                                            fake_sftp, tmp_path):
        """Half-written by a killed publish. Re-sending everything is the safe
        reading; refusing to publish is not."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        site = self._site(tmp_path)
        sftp_upload.upload_dir(site, 'test')
        (site / sftp_upload.MANIFEST_FILE).write_text('{not json')
        fake_sftp.written.clear()
        sftp_upload.upload_dir(site, 'test')
        assert len(fake_sftp.written) == 3

    def test_the_manifest_is_never_itself_uploaded(self, tournament, fake_sftp,
                                                   tmp_path):
        """It is local bookkeeping; on the server it would be a fetchable file
        describing the whole site."""
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        site = self._site(tmp_path)
        sftp_upload.upload_dir(site, 'test')
        sftp_upload.upload_dir(site, 'test')
        assert not any(sftp_upload.MANIFEST_FILE in r for r in fake_sftp.written)

    def test_progress_is_reported_per_file(self, tournament, fake_sftp, tmp_path):
        from mahj.publish import sftp_upload
        self._target(tournament['tenant'])
        seen = []
        sftp_upload.upload_dir(self._site(tmp_path), 'test',
                               progress=lambda d, t: seen.append((d, t)))
        assert seen == [(1, 3), (2, 3), (3, 3)]

    def test_no_target_is_a_noop_not_an_error(self, tournament, fake_sftp, tmp_path):
        from mahj.publish import sftp_upload
        sftp_upload.upload_dir(self._site(tmp_path), 'test')
        assert fake_sftp.written == {}

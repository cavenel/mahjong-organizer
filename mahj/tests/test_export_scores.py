"""The score half of the Excel round-trip.

One export, one import: admin_export_to_template writes the setup sheets and —
with ?scores=1 — Scores (what every seat recorded), Score sheets (every played
hand) and Standings (the ranked table, derived and ignored on the way back in).
admin_upload_from_template reads all of it back, so a workbook restores the
tournament it came from rather than an emptied copy of it.
"""
import io

from openpyxl import load_workbook

from mahj.models import Hand, Player, PublishedRound, ScoreSheet, Seat
from mahj.tests.conftest import client_for, role_user


def _sheets(content):
    """The workbook as {sheet name: [rows of values]}, header row included."""
    wb = load_workbook(io.BytesIO(content))
    return {name: [list(r) for r in wb[name].iter_rows(values_only=True)]
            for name in wb.sheetnames}


def _admin_client(tenant):
    c = client_for()
    c.force_login(role_user('exporter', tenant, admin=True))
    return c


def _upload(client, content, name='tournament.xlsx'):
    buf = io.BytesIO(content)
    buf.name = name
    return client.post('/admin_upload_from_template', {'myfile': buf})


def _importable(tenant):
    """Give every competitor an EMA-shaped id before a round-trip.

    The shared fixture uses placeholder ids ("E00001") that are not EMA numbers;
    the export writes them out faithfully and the import then rejects them, which
    says nothing about the scores these tests are about.
    """
    for i, player in enumerate(Player.objects.filter(tenant=tenant).order_by('id')):
        player.EMA_ID = f'{10000001 + i:08d}'
        player.save()


def _snapshot(tenant):
    """Everything the score tabs are supposed to carry, keyed structurally."""
    return {
        'seats': {
            (s.round_nb, s.table_nb, s.wind): (s.draw_number, s.minipoints,
                                               s.tablepoints, s.penalty)
            for s in Seat.objects.filter(tenant=tenant)
        },
        'hands': {
            (h.round_nb, h.table_nb, h.hand_nb): (h.points, h.win_by, h.win_from)
            for h in Hand.objects.filter(tenant=tenant)
        },
        'sheets': {
            (s.round_nb, s.table_nb): s.validated
            for s in ScoreSheet.objects.filter(tenant=tenant)
        },
        'published': {
            (p.round_nb, p.withheld) for p in PublishedRound.objects.filter(tenant=tenant)
        },
    }


def test_export_carries_the_scores_the_seats_and_hands_hold(tournament):
    tenant = tournament['tenant']
    resp = _admin_client(tenant).get('/admin_export_to_template?scores=1')
    assert resp.status_code == 200
    assert resp['Content-Type'] == \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert resp['Content-Disposition'] == 'attachment; filename="T_scores.xlsx"'

    sheets = _sheets(resp.content)
    assert list(sheets) == ['Options', 'Players', 'Schedule', '16 players',
                            'Scores', 'Score sheets', 'Standings']

    # --- Scores: one row per seat in the chart, scored or not.
    header, *rows = sheets['Scores']
    assert header == ['Round', 'Table', 'Wind', 'Draw #', 'Player', 'MP',
                      'Penalty', 'TP', 'Sheet']
    assert len(rows) == Seat.objects.filter(tenant=tenant).count()
    by_seat = {(r[0], r[1], r[2]): r for r in rows}
    name_of = {p.draw_number: p.full_name for p in Player.objects.filter(tenant=tenant)}
    for seat in Seat.objects.filter(tenant=tenant):
        row = by_seat[(seat.round_nb, seat.table_nb, 'ESWN'[seat.wind - 1])]
        assert row[3] == seat.draw_number
        assert row[4] == name_of[seat.draw_number]
        assert row[5] == seat.minipoints
        assert row[7] == seat.tablepoints
    # Rounds 1-2 have validated sheets; round 3 has no ScoreSheet row at all.
    assert {r[8] for r in rows if r[0] in (1, 2)} == {'validated'}
    assert {r[8] for r in rows if r[0] == 3} == {None}
    # The unscored round exports blank cells, not zeros.
    assert {r[5] for r in rows if r[0] == 3} == {None}

    # --- Score sheets: one row per played hand, winner and discarder named.
    header, *hand_rows = sheets['Score sheets']
    assert header == ['Round', 'Table', 'Hand', 'Result', 'Value',
                      'Winner wind', 'Winner', 'Dealt in wind', 'Dealt in by']
    assert len(hand_rows) == Hand.objects.filter(tenant=tenant).count()
    by_cell = {(r[0], r[1], r[2]): r for r in hand_rows}

    def seated_name(hand, wind):
        return name_of[Seat.objects.get(
            tenant=tenant, round_nb=hand.round_nb,
            table_nb=hand.table_nb, wind=wind).draw_number]

    for hand in Hand.objects.filter(tenant=tenant):
        row = by_cell[(hand.round_nb, hand.table_nb, hand.hand_nb)]
        assert row[4] == hand.points
        if hand.is_draw:
            assert row[3] == 'Draw'
            assert row[5] is None and row[6] is None      # nobody won it
        else:
            assert row[3] == 'Discard'                    # the fixture's wins all deal in
            assert row[6] == seated_name(hand, hand.win_by)
            assert row[8] == seated_name(hand, hand.win_from)

    # --- Standings: the ranked table, under a note saying it is derived.
    note, header, *st_rows = sheets['Standings']
    assert 'ignores' in note[0]
    assert header[:4] == ['Rank', 'Player', 'EMA number', 'Country']
    assert header[4:] == ['Total TP', 'Total MP'] + [
        v for r in (1, 2, 3) for v in (f'R{r} table', f'R{r} MP', f'R{r} TP')
    ]
    assert len(st_rows) == 16
    assert [r[0] for r in st_rows] == sorted(r[0] for r in st_rows)   # best first
    top = st_rows[0]
    player = Player.objects.get(tenant=tenant, full_name=top[1])
    seats = {s.round_nb: s for s in Seat.objects.filter(
        tenant=tenant, draw_number=player.draw_number)}
    assert top[5] == sum(s.minipoints for s in seats.values() if s.minipoints is not None)
    for i, round_nb in enumerate((1, 2, 3)):
        table, mp, tp = top[6 + 3 * i:9 + 3 * i]
        seat = seats[round_nb]
        assert mp == seat.minipoints
        assert tp == seat.tablepoints
        # Round 3 is seated but unscored: no score, so no table either.
        assert table == (seat.table_nb if seat.minipoints is not None else None)


def test_scores_round_trip_through_the_import(tournament):
    """Export with scores, re-import: every seat score, hand, sheet state and
    published round comes back exactly as it was."""
    tenant = tournament['tenant']
    client = _admin_client(tenant)
    _importable(tenant)
    # A penalty and a half-entered sheet, so the round-trip covers more than the
    # tidy validated case.
    Seat.objects.filter(tenant=tenant, round_nb=1, table_nb=1, wind=1).update(penalty=-30)
    ScoreSheet.objects.create(tenant=tenant, round_nb=3, table_nb=2, validated=False)
    PublishedRound.objects.filter(tenant=tenant, round_nb=2).update(withheld=True)

    before = _snapshot(tenant)
    exported = client.get('/admin_export_to_template?scores=1').content

    assert _upload(client, exported).status_code in (200, 302)
    assert _snapshot(tenant) == before


def test_an_export_without_scores_imports_an_unplayed_tournament(tournament):
    """The box unticked is the old behaviour, and the file says so: no score tabs,
    so the import loads the setup and leaves empty score sheets."""
    tenant = tournament['tenant']
    client = _admin_client(tenant)
    _importable(tenant)

    resp = client.get('/admin_export_to_template')
    assert resp['Content-Disposition'] == 'attachment; filename="T.xlsx"'
    sheets = _sheets(resp.content)
    assert list(sheets) == ['Options', 'Players', 'Schedule', '16 players']

    assert _upload(client, resp.content).status_code in (200, 302)
    assert Seat.objects.filter(tenant=tenant).count() == 48
    assert not Seat.objects.filter(tenant=tenant, minipoints__isnull=False).exists()
    assert not Hand.objects.filter(tenant=tenant).exists()
    assert not ScoreSheet.objects.filter(tenant=tenant).exists()
    assert not PublishedRound.objects.filter(tenant=tenant).exists()


def test_the_box_has_nothing_to_offer_before_play_starts(tournament):
    """Asking for scores on an unplayed tournament exports the setup alone, rather
    than three empty tabs the import would then have to read."""
    tenant = tournament['tenant']
    Hand.objects.filter(tenant=tenant).delete()
    Seat.objects.filter(tenant=tenant).update(minipoints=None, tablepoints=None)

    client = _admin_client(tenant)
    sheets = _sheets(client.get('/admin_export_to_template?scores=1').content)
    assert list(sheets) == ['Options', 'Players', 'Schedule', '16 players']

    # ... and the page shows no checkbox to ask with.
    body = client.get('/admin?page=import_template').content
    assert b'name="scores"' not in body


def test_the_page_offers_the_checkbox_once_scoring_has_started(tournament):
    body = _admin_client(tournament['tenant']).get('/admin?page=import_template').content
    assert b'name="scores"' in body


def test_riichi_export_omits_table_points(riichi_tournament):
    """Riichi ranks on minipoints alone and never fills table points in, so no
    column claims them — and the import reads the narrower sheet back."""
    tenant = riichi_tournament['tenant']
    client = _admin_client(tenant)
    _importable(tenant)
    exported = client.get('/admin_export_to_template?scores=1').content

    sheets = _sheets(exported)
    assert 'TP' not in sheets['Scores'][0]
    assert not [h for h in sheets['Standings'][1] if h and h.endswith('TP')]

    before = _snapshot(tenant)
    assert _upload(client, exported).status_code in (200, 302)
    assert _snapshot(tenant) == before


def test_teams_add_a_column_to_the_standings(tournament):
    tenant = tournament['tenant']
    settings = tournament['settings']
    settings.has_teams = True
    settings.save()
    for i, p in enumerate(Player.objects.filter(tenant=tenant).order_by('id')):
        p.team = f'Team{i % 4}'
        p.save()

    sheets = _sheets(
        _admin_client(tenant).get('/admin_export_to_template?scores=1').content)
    _note, header, *rows = sheets['Standings']
    assert header[4] == 'Team'
    assert {r[4] for r in rows} == {'Team0', 'Team1', 'Team2', 'Team3'}


def test_hand_edited_winds_and_a_reordered_sheet_still_import(tournament):
    """Columns are matched by header, not position, and winds may be spelled out —
    an organizer's edited workbook shouldn't need to mimic the export byte for byte.
    """
    tenant = tournament['tenant']
    client = _admin_client(tenant)
    _importable(tenant)
    wb = load_workbook(io.BytesIO(
        client.get('/admin_export_to_template?scores=1').content))

    scores = wb['Scores']
    # Swap the first two columns' headers and data, and spell the winds out.
    for row in range(1, scores.max_row + 1):
        a, b = scores.cell(row, 1).value, scores.cell(row, 2).value
        scores.cell(row, 1, b)
        scores.cell(row, 2, a)
        if row > 1:
            scores.cell(row, 3, {'E': 'East', 'S': 'South',
                                 'W': 'West', 'N': 'North'}[scores.cell(row, 3).value])
    buf = io.BytesIO()
    wb.save(buf)

    before = _snapshot(tenant)
    assert _upload(client, buf.getvalue()).status_code in (200, 302)
    assert _snapshot(tenant) == before


def test_an_unreadable_score_cell_is_rejected_with_nothing_changed(tournament):
    """A bad score cell is caught in the pre-check phase, so the tournament it was
    about to replace survives — the same promise a wrong file already got."""
    tenant = tournament['tenant']
    client = _admin_client(tenant)
    _importable(tenant)
    wb = load_workbook(io.BytesIO(
        client.get('/admin_export_to_template?scores=1').content))
    wb['Scores'].cell(2, 6, 'not a number')
    buf = io.BytesIO()
    wb.save(buf)

    before = _snapshot(tenant)
    resp = _upload(client, buf.getvalue())
    assert resp.status_code == 200
    assert b'nothing was changed' in resp.content
    assert _snapshot(tenant) == before


def test_scores_for_a_seat_the_chart_lacks_are_rejected(tournament):
    """The two halves of the workbook have to describe the same tournament: a score
    for an unseated place would otherwise be dropped in silence."""
    tenant = tournament['tenant']
    client = _admin_client(tenant)
    _importable(tenant)
    wb = load_workbook(io.BytesIO(
        client.get('/admin_export_to_template?scores=1').content))
    wb['Scores'].append([9, 1, 'E', 1, 'Player1 Lastname', 10, 0, 4.0, 'validated'])
    buf = io.BytesIO()
    wb.save(buf)

    resp = _upload(client, buf.getvalue())
    assert resp.status_code == 200
    assert b'does not seat' in resp.content
    # This one is caught after the deletes — import is a full replace, so it wipes
    # to empty rather than half-loading (the page says so).
    assert not Seat.objects.filter(tenant=tenant).exists()


def test_export_is_admin_only(tournament, tenant_b):
    tenant = tournament['tenant']

    anon = client_for()
    assert anon.get('/admin_export_to_template?scores=1').status_code in (302, 403)

    scorer = client_for()
    scorer.force_login(role_user('sc', tenant, scorer=True))
    assert scorer.get('/admin_export_to_template?scores=1').status_code in (302, 403)

    # An admin of another tenant has no claim on this one's scores.
    outsider = client_for()
    outsider.force_login(role_user('outsider', tenant_b, admin=True))
    assert outsider.get('/admin_export_to_template?scores=1').status_code in (302, 403)

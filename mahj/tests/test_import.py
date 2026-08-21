"""Excel-import flow (admin_upload_from_template).

Guards the round-trip the admin console runs when an organizer uploads the
tournament template: it must create one Player per player-list row (carrying the draw
number from the 'rand' column) and the full seating chart as Seat rows keyed by
draw number — with no Seat.player FK (the draw lives on Player.draw_number).
"""
import io
import json

import pytest
from django.contrib.auth.models import User
from django.test import Client
from openpyxl import load_workbook

from mahj.models import Player, Schedule, Seat, TournamentSettings, Tenant

TEMPLATE = 'mahj/static/MahjongTemplate.xlsx'


@pytest.fixture
def imp_tenant(db):
    return Tenant.objects.create(name='Import', subdomain='imp')


@pytest.fixture
def staff_client(imp_tenant):
    c = Client()
    c.defaults['HTTP_HOST'] = 'imp.example.com'  # -> subdomain 'imp'
    u = User.objects.create_superuser('imp_staff', password='pw')
    c.force_login(u)
    return c


def _filled_workbook(n=16):
    """A copy of the shipped blank template with `n` player-list rows filled in and a
    pre-assigned draw number (the 'rand' column) 1..n, so it exercises the fully
    pre-drawn import path."""
    wb = load_workbook(TEMPLATE)
    ps = wb['Players']
    for i in range(n):
        r = i + 2  # row 1 is the header
        ps.cell(r, 1, f'Last{i + 1}')     # last name
        ps.cell(r, 2, f'First{i + 1}')    # first name
        ps.cell(r, 3, 10000 + i)          # EMA id
        ps.cell(r, 4, 'Sweden')           # country
        ps.cell(r, 5, '')                 # team
        ps.cell(r, 6, i + 1)              # rand -> draw number
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'template.xlsx'
    return buf


def test_import_creates_players_and_seats(staff_client, imp_tenant):
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    assert resp.status_code in (200, 302)

    players = list(Player.objects.filter(tenant=imp_tenant))
    assert len(players) == 16
    # Every competitor carries their draw number (from the 'rand' column), unique.
    draw_numbers = sorted(p.draw_number for p in players)
    assert draw_numbers == list(range(1, 17))

    # Options sheet says 7 rounds; 16 players -> 4 tables -> 4 winds each.
    tournament = TournamentSettings.objects.get(tenant=imp_tenant)
    assert tournament.nb_rounds == 7
    seats = Seat.objects.filter(tenant=imp_tenant)
    assert seats.count() == 7 * 4 * 4  # rounds * tables * winds

    # A seat is keyed by draw number (no player FK); the competitor is the Player
    # holding that number.
    seat = seats.filter(round_nb=1, table_nb=1, wind=1).first()
    assert seat is not None
    assert Player.objects.filter(tenant=imp_tenant, draw_number=seat.draw_number).exists()
    assert not hasattr(seat, 'player_id')  # Seat has no player FK column


def test_a_cache_failure_after_a_good_import_keeps_the_tournament(
        staff_client, imp_tenant, monkeypatch):
    """The wipe-to-empty handler is deliberate policy for a *failed* import. But the
    cache bust and the display broadcast run after the transaction has committed, so a
    failure in either says nothing about the import — and while they sat inside the
    try, a Redis hiccup emptied a tournament that had just loaded cleanly, minutes
    before a round.
    """
    def boom(*a, **kw):
        raise ConnectionError('Error 111 connecting to redis:6379. Connection refused.')

    monkeypatch.setattr('mahj.views.admin_views.invalidate_leaderboard', boom)

    resp = staff_client.post('/admin_upload_from_template',
                             {'myfile': _filled_workbook(16)})
    assert resp.status_code in (200, 302)
    # The tournament is loaded, not wiped.
    assert Player.objects.filter(tenant=imp_tenant).count() == 16
    assert Seat.objects.filter(tenant=imp_tenant).count() == 7 * 4 * 4
    assert TournamentSettings.objects.get(tenant=imp_tenant).nb_rounds == 7


def test_a_broadcast_failure_after_a_good_import_keeps_the_tournament(
        staff_client, imp_tenant, monkeypatch):
    """Same for the other post-commit call."""
    def boom(*a, **kw):
        raise RuntimeError('channel layer unavailable')

    monkeypatch.setattr('mahj.views.admin_views.broadcast_publish_state', boom)

    resp = staff_client.post('/admin_upload_from_template',
                             {'myfile': _filled_workbook(16)})
    assert resp.status_code in (200, 302)
    assert Player.objects.filter(tenant=imp_tenant).count() == 16


def _snapshot(tenant):
    players = {
        (p.full_name, p.first_name, p.last_name, p.EMA_ID, p.country, p.team,
         p.draw_number)
        for p in Player.objects.filter(tenant=tenant)
    }
    seats = {
        (s.round_nb, s.table_nb, s.wind, s.draw_number)
        for s in Seat.objects.filter(tenant=tenant)
    }
    schedule = {
        (i.day, i.time, i.name, i.is_round)
        for i in Schedule.objects.filter(tenant=tenant)
    }
    return players, seats, schedule


def test_export_round_trips_through_import(staff_client, imp_tenant):
    """Import a tournament, add a per-player team and a schedule, export via
    admin_export_to_template, then re-import: the player list (incl. team), the full
    seating chart and the schedule (incl. is_round) must match."""
    staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})

    # Populate fields beyond name/EMA/draw so the round-trip is actually exercised.
    for i, p in enumerate(Player.objects.filter(tenant=imp_tenant).order_by('id')):
        p.team = f'Team{i % 4}'
        p.save()

    before = _snapshot(imp_tenant)

    resp = staff_client.get('/admin_export_to_template')
    assert resp.status_code == 200
    assert resp['Content-Type'] == \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # A single seating sheet for this field size, not the template's full set.
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ['Options', 'Players', 'Schedule', '16 players']

    exported = io.BytesIO(resp.content)
    exported.seek(0)
    exported.name = 'template.xlsx'

    resp = staff_client.post('/admin_upload_from_template', {'myfile': exported})
    assert resp.status_code in (200, 302)

    assert _snapshot(imp_tenant) == before


def test_import_without_rand_leaves_players_undrawn(staff_client, imp_tenant):
    """No 'rand' column -> player list imported but not yet drawn (draw_number NULL);
    the seating chart still exists, to be filled by randomize/team-draw."""
    wb = load_workbook(TEMPLATE)
    ps = wb['Players']
    for i in range(16):
        r = i + 2
        ps.cell(r, 1, f'Last{i + 1}')
        ps.cell(r, 2, f'First{i + 1}')
        ps.cell(r, 3, 20000 + i)
        ps.cell(r, 4, 'France')
        ps.cell(r, 5, '')
        # no rand
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'template.xlsx'

    resp = staff_client.post('/admin_upload_from_template', {'myfile': buf})
    assert resp.status_code in (200, 302)

    assert Player.objects.filter(tenant=imp_tenant).count() == 16
    assert Player.objects.filter(tenant=imp_tenant, draw_number__isnull=True).count() == 16
    # Seats still created (structural chart), just not linked to a competitor yet.
    assert Seat.objects.filter(tenant=imp_tenant).count() == 7 * 4 * 4


# --- robustness / hardening (PR A: F-C1, F-H9, F-M11, F-M12) ---------------

_KEEP = object()  # sentinel: leave the template's default number of rounds


def _workbook(rows, nb_rounds=_KEEP):
    """Build an importable workbook from `rows` — a list of
    (last, first, ema, country, team, rand) tuples written into the Players sheet
    (a None cell is left blank; an all-None tuple is a blank spacer row). Pass
    `nb_rounds` (incl. None for a blank cell) to override the Options 'number of
    rounds' cell (B3); omit it to keep the template default."""
    wb = load_workbook(TEMPLATE)
    ps = wb['Players']
    for i, cells in enumerate(rows):
        for c, val in enumerate(cells, start=1):
            if val is not None:
                ps.cell(row=i + 2, column=c, value=val)
    if nb_rounds is not _KEEP:
        # Assign .value directly: cell(..., value=None) is a no-op in openpyxl.
        wb['Options'].cell(row=3, column=2).value = nb_rounds
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'template.xlsx'
    return buf


def _rows16(**overrides):
    """16 complete players, draw numbers 1..16. `overrides` is {row_index: tuple}
    to replace individual rows (0-based over the 16)."""
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', None, i + 1)
            for i in range(16)]
    for idx, tup in overrides.items():
        rows[idx] = tup
    return rows


def test_blank_row_midlist_does_not_truncate(staff_client, imp_tenant):
    """F-H9: a fully-blank spacer row in the middle of the player list is skipped, not
    treated as the end of the list (which used to drop everyone below it)."""
    rows = _rows16()
    rows.insert(8, (None, None, None, None, None, None))  # blank spacer mid-list
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code in (200, 302)
    assert Player.objects.filter(tenant=imp_tenant).count() == 16


def test_team_whitespace_collapsed_and_numeric_cell(staff_client, imp_tenant):
    """F-M11 + F-M12a: internal whitespace is collapsed so "Team  A" and "Team A"
    are one team; a numeric team cell is coerced to a string instead of crashing;
    case is NOT folded ("sweden" and "Sweden" stay two distinct teams). Every team
    here is a valid group of four (see test_uneven_team_rejected)."""
    teams = ['Team  A'] * 2 + ['Team A'] * 2 + [123] * 4 + ['sweden'] * 4 + ['Sweden'] * 4
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', teams[i], i + 1)
            for i in range(16)]
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code in (200, 302)
    stored = set(Player.objects.filter(tenant=imp_tenant).values_list('team', flat=True))
    assert stored == {'Team A', '123', 'sweden', 'Sweden'}
    # The two whitespace variants collapsed into one four-person "Team A".
    assert Player.objects.filter(tenant=imp_tenant, team='Team A').count() == 4


def test_uneven_team_rejected(staff_client, imp_tenant):
    """F-M11: teams are always groups of four, so a team of a different size is
    rejected (and per F-C1 the import leaves the tournament empty). This is what
    surfaces a typo/case split like "Sweden"(3) vs "sweden"(1)."""
    teams = ['Sweden'] * 3 + ['sweden'] * 1 + ['Norway'] * 4 + ['Denmark'] * 4 + ['Finland'] * 4
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', teams[i], i + 1)
            for i in range(16)]
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 0


def test_names_stored_raw_from_two_columns(staff_client, imp_tenant):
    """The Last/First columns are stored raw (mixed case preserved, no title-casing)
    and full_name is the "First Last" join. Covers a mononym (blank surname), a
    multi-word surname kept whole, and casing that .title() used to mangle."""
    rows = _rows16()
    rows[0] = ('', 'Cher', 90000, 'France', None, 1)               # mononym
    rows[1] = ('Van Der Berg', 'Chris', 90001, 'Sweden', None, 2)  # multi-word surname
    rows[2] = ('McDonald', 'chris', 90002, 'Scotland', None, 3)    # casing preserved
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code in (200, 302)

    cher = Player.objects.get(tenant=imp_tenant, draw_number=1)
    assert (cher.first_name, cher.last_name, cher.full_name) == ('Cher', '', 'Cher')

    berg = Player.objects.get(tenant=imp_tenant, draw_number=2)
    assert berg.last_name == 'Van Der Berg'
    assert berg.full_name == 'Chris Van Der Berg'

    mac = Player.objects.get(tenant=imp_tenant, draw_number=3)
    assert mac.last_name == 'McDonald'  # not "Mcdonald"
    assert mac.full_name == 'chris McDonald'


def test_short_name_disambiguates_shared_first_name(staff_client, imp_tenant):
    """short_name is the bare first name when unique, else first name + the shortest
    surname prefix that separates same-first-name competitors ("Chris D.", growing
    to "Chris Dere." when two share a prefix)."""
    rows = _rows16()
    rows[0] = ('Derek', 'Chris', 90000, 'Sweden', None, 1)
    rows[1] = ('Dupont', 'Chris', 90001, 'Sweden', None, 2)
    rows[2] = ('Dervinson', 'Chris', 90002, 'Sweden', None, 3)
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code in (200, 302)

    def short(dn):
        return Player.objects.get(tenant=imp_tenant, draw_number=dn).short_name

    # A unique first name stays bare.
    assert short(4) == 'First4'
    # Three Chrises: Dupont splits on the first letter; Derek/Dervinson need "Dere"/"Derv".
    assert short(1) == 'Chris Dere.'
    assert short(2) == 'Chris Du.'
    assert short(3) == 'Chris Derv.'


def test_names_round_trip_through_export(staff_client, imp_tenant):
    """A mononym and a multi-word surname survive export -> re-import byte-identical
    (they no longer depend on re-splitting full_name)."""
    rows = _rows16()
    rows[0] = ('', 'Cher', 90000, 'France', None, 1)
    rows[1] = ('Van Der Berg', 'Chris', 90001, 'Sweden', None, 2)
    staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    before = _snapshot(imp_tenant)

    resp = staff_client.get('/admin_export_to_template')
    assert resp.status_code == 200
    exported = io.BytesIO(resp.content)
    exported.seek(0)
    exported.name = 'template.xlsx'
    staff_client.post('/admin_upload_from_template', {'myfile': exported})

    assert _snapshot(imp_tenant) == before


def test_absent_ema_left_blank(staff_client, imp_tenant):
    """F-M12c: a genuinely absent EMA id is silently blank (most players have none)."""
    rows = [(f'Last{i + 1}', f'First{i + 1}', None, 'Sweden', None, i + 1) for i in range(16)]
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code in (200, 302)
    players = Player.objects.filter(tenant=imp_tenant)
    assert players.count() == 16
    assert all(p.EMA_ID == '' for p in players)


def test_invalid_ema_fails_whole_import(staff_client, imp_tenant):
    """F-M12c: a present-but-unparseable EMA id fails the import (and per F-C1
    leaves the tournament empty) rather than silently blanking the id."""
    rows = _rows16()
    rows[4] = ('LastX', 'FirstX', 'not-a-number', 'Sweden', None, 5)
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 0


def test_zero_draw_number_rejected(staff_client, imp_tenant):
    """F-M12d: draw number 0 collides with the empty-seat sentinel, so it is
    rejected (draw numbers are 1-based)."""
    rows = _rows16()
    rows[0] = ('Last1', 'First1', 90000, 'Sweden', None, 0)
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert resp.status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 0


def test_blank_rounds_rejected(staff_client, imp_tenant):
    """F-M12b: a blank/zero rounds count creates no seating and used to 'succeed'
    with nothing playable; it is now rejected."""
    resp = staff_client.post(
        '/admin_upload_from_template', {'myfile': _workbook(_rows16(), nb_rounds=None)})
    assert resp.status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 0
    assert Seat.objects.filter(tenant=imp_tenant).count() == 0


class TestPrecheckRejectsUntouched:
    """The checks that catch the realistic wrong-file mistakes (missing sheets,
    unreadable rounds count, unplayable player count, colliding draw numbers,
    not-a-workbook) run BEFORE the first delete(): a workbook failing one is
    rejected with the tournament untouched — that untouched-ness is the property
    under test. A workbook that passes them and fails deeper in the parse keeps
    the documented wipe-to-empty (see test_failed_import_leaves_tournament_empty)."""

    @pytest.fixture
    def live_before(self, staff_client, imp_tenant):
        """A populated tournament, and its snapshot to compare against."""
        staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
        assert Player.objects.filter(tenant=imp_tenant).count() == 16
        return _snapshot(imp_tenant)

    def _assert_rejected_untouched(self, resp, imp_tenant, live_before):
        assert resp.status_code == 200
        assert 'nothing was changed' in resp.content.decode()
        assert _snapshot(imp_tenant) == live_before

    def test_missing_required_sheet(self, staff_client, imp_tenant, live_before):
        wb = load_workbook(TEMPLATE)
        wb.remove(wb['Players'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'template.xlsx'
        resp = staff_client.post('/admin_upload_from_template', {'myfile': buf})
        self._assert_rejected_untouched(resp, imp_tenant, live_before)
        assert 'Players' in resp.content.decode()  # the message names the sheet

    def test_unreadable_rounds_count(self, staff_client, imp_tenant, live_before):
        resp = staff_client.post(
            '/admin_upload_from_template',
            {'myfile': _workbook(_rows16(), nb_rounds=None)})
        self._assert_rejected_untouched(resp, imp_tenant, live_before)

    def test_player_count_not_a_multiple_of_four(self, staff_client, imp_tenant,
                                                 live_before):
        resp = staff_client.post(
            '/admin_upload_from_template', {'myfile': _workbook(_rows16()[:15])})
        self._assert_rejected_untouched(resp, imp_tenant, live_before)
        assert '15 competitors' in resp.content.decode()

    def test_empty_player_list(self, staff_client, imp_tenant, live_before):
        resp = staff_client.post(
            '/admin_upload_from_template', {'myfile': _workbook([])})
        self._assert_rejected_untouched(resp, imp_tenant, live_before)

    def test_duplicate_draw_numbers(self, staff_client, imp_tenant, live_before):
        """Two competitors sharing a 'rand' value used to hit the per-tenant
        unique constraint mid-load — a traceback and a wiped tournament."""
        rows = _rows16()
        rows[7] = ('Last8', 'First8', 90007, 'Sweden', None, 5)  # 5 already taken
        resp = staff_client.post(
            '/admin_upload_from_template', {'myfile': _workbook(rows)})
        self._assert_rejected_untouched(resp, imp_tenant, live_before)
        assert 'Draw number 5' in resp.content.decode()

    def test_not_a_workbook(self, staff_client, imp_tenant, live_before):
        buf = io.BytesIO(b'this is not an xlsx file')
        buf.name = 'notes.txt'
        resp = staff_client.post('/admin_upload_from_template', {'myfile': buf})
        self._assert_rejected_untouched(resp, imp_tenant, live_before)

    def test_a_good_workbook_still_imports(self, staff_client, imp_tenant, live_before):
        """The pre-checks must not reject what the app itself produces: a fresh
        import over the live tournament replaces it wholesale."""
        rows = [(f'New{i + 1}', f'Player{i + 1}', 80000 + i, 'Norway', None, i + 1)
                for i in range(16)]
        resp = staff_client.post(
            '/admin_upload_from_template', {'myfile': _workbook(rows)})
        assert resp.status_code in (200, 302)
        players = Player.objects.filter(tenant=imp_tenant)
        assert players.count() == 16
        assert set(players.values_list('country', flat=True)) == {'Norway'}


def test_failed_import_leaves_tournament_empty(staff_client, imp_tenant):
    """F-C1: a failed re-import over a live tournament wipes it to a clean empty
    state (no partial/ghost tournament, and not silently reverted to the old one)."""
    staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    assert Player.objects.filter(tenant=imp_tenant).count() == 16

    rows = _rows16()
    rows[3] = ('LastBad', 'FirstBad', 'garbage', 'Sweden', None, 4)  # bad EMA -> fails
    staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})

    assert Player.objects.filter(tenant=imp_tenant).count() == 0
    assert Seat.objects.filter(tenant=imp_tenant).count() == 0
    ts = TournamentSettings.objects.get(tenant=imp_tenant)
    assert ts.nb_rounds == 0
    assert ts.fullname == ''


def test_import_evaluates_seating_formulas(staff_client, imp_tenant):
    """The shipped 28-player sheet mirrors its high half with formulas ("=14+8"
    -> draw 22). openpyxl drops Excel's cached formula results whenever the
    workbook is re-saved (as _filled_workbook does), so the importer must
    evaluate the formulas — otherwise half of every table would import as empty
    seats (draw number 0). Every seat must carry a real draw number 1..28."""
    resp = staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(28)})
    assert resp.status_code in (200, 302)

    seats = Seat.objects.filter(tenant=imp_tenant)
    assert seats.count() == 7 * 7 * 4  # rounds * tables * winds
    assert not seats.filter(draw_number=0).exists()  # no un-evaluated formula gaps
    for r in range(1, 8):
        drawn = sorted(s.draw_number for s in seats.filter(round_nb=r))
        assert drawn == list(range(1, 29))  # each competitor seated once per round


def test_broken_seating_chart_rejected(staff_client, imp_tenant):
    """A seating sheet that doesn't seat every competitor exactly once per round
    (here a blanked cell) is rejected, leaving the tournament empty — rather than
    silently loading a chart with a ghost empty seat."""
    wb = load_workbook(TEMPLATE)
    ps = wb['Players']
    for i in range(16):
        r = i + 2
        ps.cell(r, 1, f'Last{i + 1}')
        ps.cell(r, 2, f'First{i + 1}')
        ps.cell(r, 3, 90000 + i)
        ps.cell(r, 4, 'Sweden')
        ps.cell(r, 6, i + 1)
    wb['16 players'].cell(row=3, column=2).value = None  # blank one seat in round 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'template.xlsx'

    resp = staff_client.post('/admin_upload_from_template', {'myfile': buf})
    assert resp.status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 0
    assert Seat.objects.filter(tenant=imp_tenant).count() == 0


# --- in-app seating generation (admin_generate_seating) --------------------

def test_generate_seating_builds_chart_and_keeps_players(staff_client, imp_tenant):
    """Generating replaces the seating chart (a full chart for every round) and
    returns quality measures, while keeping the player list and draw."""
    staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    draws_before = sorted(Player.objects.filter(tenant=imp_tenant)
                          .values_list('draw_number', flat=True))

    # No body -> auto method, apply immediately (an empty test-client post would
    # otherwise encode a non-empty multipart body, which the JSON contract 400s).
    resp = staff_client.post('/admin_generate_seating', content_type='application/json')
    assert resp.status_code == 200
    data = resp.json()
    assert data['ok'] is True
    assert data['measures']['all_seated'] is True

    assert Seat.objects.filter(tenant=imp_tenant).count() == 7 * 4 * 4  # rounds*tables*winds
    # Player list and draw are untouched — the chart is independent of the people.
    assert Player.objects.filter(tenant=imp_tenant).count() == 16
    assert sorted(Player.objects.filter(tenant=imp_tenant)
                  .values_list('draw_number', flat=True)) == draws_before
    for r in range(1, 8):
        drawn = sorted(s.draw_number for s in Seat.objects.filter(tenant=imp_tenant, round_nb=r))
        assert drawn == list(range(1, 17))


def test_generate_seating_algebraic_infeasible_refuses_without_touching_chart(staff_client, imp_tenant):
    """Explicitly requesting the deterministic method where no rematch-free chart
    exists is refused (400) and the existing chart is left intact. (The default
    'auto'/best-effort still produces a chart — see the best-effort test.)"""
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', f'Team{i // 4}', i + 1)
            for i in range(16)]  # 4 teams / 7 rounds -> algebraic infeasible
    staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    before = Seat.objects.filter(tenant=imp_tenant).count()
    assert before == 7 * 4 * 4

    resp = staff_client.post('/admin_generate_seating',
                             data=json.dumps({'method': 'algebraic'}),
                             content_type='application/json')
    assert resp.status_code == 400
    assert 'error' in resp.json()
    assert Seat.objects.filter(tenant=imp_tenant).count() == before  # untouched


def test_import_without_seating_sheet_keeps_players_seatless(staff_client, imp_tenant):
    """A workbook with no '<N> players' seating sheet imports the player list/schedule
    and leaves the tournament without a chart (to be generated later on the Seating
    page) instead of failing the whole import."""
    wb = load_workbook(TEMPLATE)
    ps = wb['Players']
    for i in range(16):
        r = i + 2
        ps.cell(r, 1, f'Last{i + 1}')
        ps.cell(r, 2, f'First{i + 1}')
        ps.cell(r, 3, 90000 + i)
        ps.cell(r, 4, 'Sweden')
        ps.cell(r, 6, i + 1)
    for name in list(wb.sheetnames):
        if name.endswith(' players'):
            del wb[name]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'template.xlsx'

    resp = staff_client.post('/admin_upload_from_template', {'myfile': buf})
    assert resp.status_code in (200, 302)
    assert Player.objects.filter(tenant=imp_tenant).count() == 16      # player list imported
    assert Seat.objects.filter(tenant=imp_tenant).count() == 0         # no chart yet

    # ...and a chart can then be generated for it.
    resp = staff_client.post('/admin_generate_seating', content_type='application/json')
    assert resp.status_code == 200
    assert Seat.objects.filter(tenant=imp_tenant).count() == 7 * 4 * 4


def test_generate_seating_preview_does_not_write(staff_client, imp_tenant):
    """A preview (apply=false) returns measures without changing the stored chart."""
    staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    before = {(s.round_nb, s.table_nb, s.wind, s.draw_number)
              for s in Seat.objects.filter(tenant=imp_tenant)}

    resp = staff_client.post('/admin_generate_seating',
                             data=json.dumps({'method': 'greedy', 'seed': 3, 'apply': False}),
                             content_type='application/json')
    assert resp.status_code == 200
    data = resp.json()
    assert data['ok'] is True and data['applied'] is False
    assert 'measures' in data
    after = {(s.round_nb, s.table_nb, s.wind, s.draw_number)
             for s in Seat.objects.filter(tenant=imp_tenant)}
    assert after == before  # nothing written


def test_seating_page_renders(staff_client, imp_tenant):
    """The Seating admin page renders, showing the generate controls."""
    staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    resp = staff_client.get('/admin?page=seating')
    assert resp.status_code == 200
    html = resp.content.decode()
    assert 'Generate a new seating' in html
    assert 'Current seating' in html
    assert 'cross_positions' in html  # print link shown once a chart exists
    assert 'Number of tries' in html  # best-effort search controls are visible
    assert 'Variation (seed)' in html


def test_cross_positions_before_draw_uses_placeholder(staff_client, imp_tenant):
    """The cross-position sheet renders before any draw (no 500), labelling
    unclaimed draw slots 'Player N' instead of dereferencing a missing player."""
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', None, None)
            for i in range(16)]  # no 'rand' -> players undrawn, but the chart exists
    staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    assert Seat.objects.filter(tenant=imp_tenant).exists()
    assert Player.objects.filter(tenant=imp_tenant, draw_number__isnull=True).count() == 16

    resp = staff_client.get('/cross_positions')
    assert resp.status_code == 200
    assert 'Player 1' in resp.content.decode()
    # Team variant must also not crash when nobody is drawn in.
    resp = staff_client.get('/cross_positions?per_team=1')
    assert resp.status_code == 200


def test_dashboard_reports_seating_status(staff_client, imp_tenant):
    """The dashboard's setup checklist shows whether a seating chart exists."""
    # No import yet -> no seating.
    resp = staff_client.get('/admin?page=welcome')
    assert 'No seating chart' in resp.content.decode()
    # After importing a chart -> ready.
    staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    resp = staff_client.get('/admin?page=welcome')
    assert 'Seating chart ready' in resp.content.decode()


def test_seating_page_team_infeasible_offers_best_effort(staff_client, imp_tenant):
    """For a team field where a perfect chart is impossible, the page still offers
    the best-effort method (with a teammate-clash caveat) rather than refusing."""
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', f'Team{i // 4}', i + 1)
            for i in range(16)]  # 4 teams / 7 rounds -> algebraic infeasible
    staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    resp = staff_client.get('/admin?page=seating')
    assert resp.status_code == 200
    html = resp.content.decode()
    assert 'Best-effort search' in html
    assert 'minimises teammate clashes' in html
    assert 'rematch-free' in html  # deterministic shown disabled with a reason


def test_generate_seating_team_best_effort_when_infeasible(staff_client, imp_tenant):
    """Generating a team chart where a perfect one is impossible succeeds with the
    best-effort search and writes a full chart, rather than refusing."""
    rows = [(f'Last{i + 1}', f'First{i + 1}', 90000 + i, 'Sweden', f'Team{i // 4}', i + 1)
            for i in range(16)]
    staff_client.post('/admin_upload_from_template', {'myfile': _workbook(rows)})
    resp = staff_client.post('/admin_generate_seating',
                             data=json.dumps({'method': 'greedy', 'apply': True}),
                             content_type='application/json')
    assert resp.status_code == 200
    assert resp.json()['ok'] is True
    assert Seat.objects.filter(tenant=imp_tenant).count() == 7 * 4 * 4


# --------------------------------------------------------------------------
# Tenancy and atomicity of the import itself (F10 + import atomicity)
# --------------------------------------------------------------------------

def test_import_reads_the_upload_not_a_shared_path(staff_client, imp_tenant, tmp_path):
    """F10: every import used to be staged at one fixed path, BASE_DIR/tmp/template.xlsx.
    Two tenants importing at once raced over it, and the loser could load the other's
    workbook — after deleting its own players. The upload is now read directly, so no
    shared file is written at all."""
    from mahj.views.helpers import BASE_DIR
    staged = BASE_DIR / 'tmp' / 'template.xlsx'
    before = staged.stat().st_mtime if staged.exists() else None

    resp = staff_client.post('/admin_upload_from_template', {'myfile': _filled_workbook(16)})
    assert resp.status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 16

    after = staged.stat().st_mtime if staged.exists() else None
    assert after == before, 'the import still writes a shared staging file'


def test_two_tenants_importing_do_not_cross(staff_client, imp_tenant, db):
    """The other half of F10: each tenant ends up with exactly its own player list.
    Sequential here — the shared-path race is what the test above rules out — but this
    pins the tenant scoping of the import itself."""
    other = Tenant.objects.create(name='Other import', subdomain='imp2')
    other_client = Client()
    other_client.defaults['HTTP_HOST'] = 'imp2.example.com'
    u = User.objects.create_superuser('imp2_staff', password='pw')
    other_client.force_login(u)

    assert staff_client.post(
        '/admin_upload_from_template', {'myfile': _filled_workbook(16)}).status_code == 200
    assert other_client.post(
        '/admin_upload_from_template', {'myfile': _filled_workbook(20)}).status_code == 200

    assert Player.objects.filter(tenant=imp_tenant).count() == 16
    assert Player.objects.filter(tenant=other).count() == 20
    # And no seat from one tenant leaked into the other's chart.
    assert Seat.objects.filter(tenant=imp_tenant).exclude(draw_number__lte=16).count() == 0


def test_a_failure_mid_import_leaves_nothing_behind(staff_client, imp_tenant):
    """The import is one transaction, and a validation failure mid-parse wipes to
    empty by design — never a half-loaded tournament, and never the old one
    silently restored. The trigger must be a *deep* failure (here a broken
    seating chart, found after players and schedule are already written): the
    cheap mistakes are pre-checked before anything is deleted and reject with
    the tournament untouched instead (see TestPrecheckRejectsUntouched)."""
    assert staff_client.post(
        '/admin_upload_from_template', {'myfile': _filled_workbook(16)}).status_code == 200
    assert Player.objects.filter(tenant=imp_tenant).count() == 16

    wb = load_workbook(TEMPLATE)
    ps = wb['Players']
    for i in range(16):
        r = i + 2
        ps.cell(r, 1, f'Last{i + 1}')
        ps.cell(r, 2, f'First{i + 1}')
        ps.cell(r, 6, i + 1)
    wb['16 players'].cell(row=3, column=2).value = None  # blank one round-1 seat
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'broken.xlsx'

    resp = staff_client.post('/admin_upload_from_template', {'myfile': buf})
    assert resp.status_code == 200          # the page renders with an error banner
    assert Player.objects.filter(tenant=imp_tenant).count() == 0
    assert Seat.objects.filter(tenant=imp_tenant).count() == 0
    assert Schedule.objects.filter(tenant=imp_tenant).count() == 0

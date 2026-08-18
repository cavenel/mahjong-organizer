"""Public `/` landing page: the tournament desktop view."""
import io

from django.contrib.auth import logout
from django.core.cache import cache
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..models import Hand, ScoreSheet, Schedule, Seat
from ..scoring import _attach_players, team_standings
from .helpers import can_access_admin, get_tenant, get_tournament, is_tenant_admin
from .scoring import (
    LEADERBOARD_TTL, scores_per_player_json, stat_all_rounds, stat_rounds,
    stats_export, table_stats, table_stats_rounds, tournament_seating,
)

HTML_CACHE_TTL = LEADERBOARD_TTL  # same TTL as data; also invalidated by signals


def desktop(request):
    tenant = get_tenant(request)
    if tenant is None:
        return HttpResponseRedirect('admin')

    if request.GET.get('logout') == "1":
        logout(request)

    subdomain = tenant.subdomain if tenant else ''
    # "admin" here = full view of this tenant (reveals withheld/unpublished scores):
    # a tenant admin or a platform superuser. Tenant-scoped, so a member of another
    # tenant on this subdomain is treated as a plain viewer.
    is_admin = is_tenant_admin(request)
    full_view = is_admin

    # The overflow menu (Admin / Log out / Login) varies by role, but the page is
    # cached as one HTML blob, so the cache key must capture every menu variant or
    # the first render leaks its menu to the whole bucket. One token covers both
    # the data view (is_admin → full_view) and the menu: the anonymous crowd all
    # share 'anon' (so nginx still microcaches a single `/` entry), while the few
    # privileged sessions each get their own variant.
    authenticated = request.user.is_authenticated
    user_can_access_admin = can_access_admin(request)
    view = ('staff' if is_admin else 'admin' if user_can_access_admin
            else 'user' if authenticated else 'anon')

    # A static-export render has no auth on the target host, so it drops the
    # login/admin menu — and must neither read nor write the live anon HTML cache
    # (that variant still needs the menu).
    static_export = getattr(request, '_static_export', False)

    # Cache the full rendered HTML — it only changes when scores/tournament are written,
    # both of which already call invalidate_leaderboard() which deletes this key.
    html_key = f'desktop_html:{subdomain}:{view}'
    if not static_export:
        cached_html = cache.get(html_key)
        if cached_html is not None:
            return HttpResponse(cached_html)

    tournament = get_tournament(request)
    nb_rounds = tournament.nb_rounds

    # Fetch positions and hands once; share between standings, seating, and stats.
    positions = _attach_players(tenant, list(
        Seat.objects.filter(tenant=tenant).order_by('round_nb')
    ))
    hands = list(Hand.objects.filter(tenant=tenant))
    valid_pairs = set(
        ScoreSheet.objects.filter(tenant=tenant, validated=True)
        .values_list('round_nb', 'table_nb')
    )
    scores_json = scores_per_player_json(request, full_view=is_admin, positions=positions)
    seating, player_table = tournament_seating(
        request, full_view=full_view, valid_pairs=valid_pairs, positions=positions,
    )
    seating_json = [
        {
            'round_nb': r['round_nb'],
            'tables': [
                {
                    'table_nb': t['table_nb'],
                    'has_scores': t.get('has_scores', False),
                    'seats': [
                        {
                            'wind': s['wind'],
                            'mp': s['mp'],
                            'tp': s['tp'],
                            'name': s['name'],
                            'player': (
                                {'id': s['player'].id, 'full_name': s['player'].full_name}
                                if s['player'] else None
                            ),
                        }
                        for s in t['seats']
                    ],
                }
                for t in r['tables']
            ],
        }
        for r in seating
    ]

    rows = []
    for s in scores_json:
        tables = [player_table.get((s['player_id'], r)) for r in range(1, nb_rounds + 1)]
        scores_with_table = []
        for r_idx, sc in enumerate(s.get('scores', [])):
            scores_with_table.append({
                'tp': sc.get('tp'),
                'mp': sc.get('mp'),
                'table_nb': tables[r_idx] if r_idx < len(tables) else None,
                'round_nb': r_idx + 1,
            })
        rows.append({
            'player_id': s['player_id'],
            'name': s['name'],
            'flag': s['flag'],
            'country': s.get('country', ''),
            'pos': s['pos'],
            'pos_se': s.get('pos_se'),
            'total': s['total'],
            'team': s.get('team', ''),
            'scores': scores_with_table,
        })

    uses_teams = tournament.has_teams
    team_rows = team_standings(rows, tournament, nb_rounds) if uses_teams else []

    schedule_key = f'schedule:{subdomain}'
    schedule = cache.get(schedule_key)
    if schedule is None:
        schedule = list(Schedule.objects.filter(tenant=tenant).order_by('id'))
        cache.set(schedule_key, schedule, 300)
    stat_rounds_data = stat_rounds(request, full_view=full_view, positions=positions, hands=hands)
    stat_all_data = stat_all_rounds(request, full_view=full_view, positions=positions, hands=hands)
    stat_tables_data = table_stats(request, full_view=full_view, positions=positions, hands=hands)
    stat_tables_rounds_data = table_stats_rounds(request, full_view=full_view, positions=positions, hands=hands)
    # Pair each round's winner stats with its table stats so one per-round loop in the
    # template can render both panels (both lists are round-aligned, length round_max).
    stat_rounds_combined = [
        {'winners': w, 'tables': t}
        for w, t in zip(stat_rounds_data, stat_tables_rounds_data)
    ]

    context = {
        'tournament': tournament,
        'rows': rows,
        'rounds': list(range(1, 1 + nb_rounds)),
        'max_round': nb_rounds,
        'seating': seating,
        'seating_json': seating_json,
        'tenant': tenant,
        'schedule': schedule,
        'stat_rounds': stat_rounds_data,
        'stat_all': stat_all_data,
        'stat_tables': stat_tables_data,
        'stat_rounds_combined': stat_rounds_combined,
        'uses_teams': uses_teams,
        'team_rows': team_rows,
        'user_authenticated': authenticated,
        'user_can_access_admin': user_can_access_admin,
        'static_export': static_export,
    }
    template = loader.get_template('mahj/desktop.html')
    html = template.render(context, request)
    if not static_export:
        cache.set(html_key, html, HTML_CACHE_TTL)
    return HttpResponse(html)


# Excel color-scale endpoints: red (worst) → yellow → green (best). 'down' columns
# (rank, deal-ins, losses) are where a *lower* value is better, so their scale is
# flipped so green still marks the good end.
_XLSX_RED, _XLSX_MID, _XLSX_GREEN = 'F8696B', 'FFEB84', '63BE7B'


def _color_scale(good='up'):
    lo, hi = (_XLSX_RED, _XLSX_GREEN) if good == 'up' else (_XLSX_GREEN, _XLSX_RED)
    return ColorScaleRule(
        start_type='min', start_color=lo,
        mid_type='percentile', mid_value=50, mid_color=_XLSX_MID,
        end_type='max', end_color=hi,
    )


def stats_xlsx(request):
    """Download every displayed per-player stat as one colour-scaled Excel sheet.

    Same visibility as the stats tab: public viewers get published rounds, admins
    get every scored round. Each numeric column carries a red→green colour scale
    (green = the better end), matching the on-screen player/tournament stats.
    """
    tenant = get_tenant(request)
    if tenant is None:
        return HttpResponseRedirect('admin')
    subdomain = tenant.subdomain if tenant else ''
    is_admin = is_tenant_admin(request)
    full_view = is_admin

    positions = _attach_players(tenant, list(
        Seat.objects.filter(tenant=tenant).order_by('round_nb')
    ))
    hands = list(Hand.objects.filter(tenant=tenant))
    data = stats_export(request, full_view=full_view, positions=positions, hands=hands)

    is_mcr = data['rules'] == 'MCR'
    uses_teams = data['uses_teams']
    nb_rounds = data['round_max']

    # (header, row-accessor, colour direction | None). Per-round columns are spliced
    # in after the totals; everything after is a stat straight from stats_export.
    cols = [
        ('Rank', lambda r: r['rank'], 'down'),
        ('Player', lambda r: r['name'], None),
        ('Country', lambda r: r['country'], None),
    ]
    if uses_teams:
        cols.append(('Team', lambda r: r['team'], None))
    if is_mcr:
        cols.append(('Total TP', lambda r: r['total_tp'], 'up'))
    cols.append(('Total MP', lambda r: r['total_mp'], 'up'))
    for i in range(nb_rounds):
        cols.append((f'R{i + 1} MP', (lambda i: lambda r: (r['scores'][i]['mp'] if i < len(r['scores']) else None))(i), 'up'))
        if is_mcr:
            cols.append((f'R{i + 1} TP', (lambda i: lambda r: (r['scores'][i]['tp'] if i < len(r['scores']) else None))(i), 'up'))
    cols += [
        ('1st', lambda r: r['placement'][1], 'up'),
        ('2nd', lambda r: r['placement'][2], 'up'),
        ('3rd', lambda r: r['placement'][3], 'down'),
        ('4th', lambda r: r['placement'][4], 'down'),
        ('Hands', lambda r: r['total_hands'], 'up'),
        ('Wins', lambda r: r['wins'], 'up'),
        ('Win self-draw', lambda r: r['sd_win'], 'up'),
        ('Win discard', lambda r: r['ron_win'], 'up'),
        ('Self-draw % of wins', lambda r: r['sd_win_share_pct'], 'up'),
        ('Self-draw rate %', lambda r: r['sd_rate_pct'], 'up'),
        ('Avg hand value', lambda r: r['avg_hand_value'], 'up'),
        ('Biggest hand', lambda r: r['biggest_hand'], 'up'),
        ('Deal-ins', lambda r: r['deal_in'], 'down'),
        ('Deal-in %', lambda r: r['deal_in_pct'], 'down'),
        ('Self-draw victim', lambda r: r['sd_lose'], 'down'),
        ('Self-draw victim %', lambda r: r['sd_lose_pct'], 'down'),
        ('Draws', lambda r: r['draws'], 'up'),
        ("Opp strength total", lambda r: r['opp_total'], 'up'),
        ('Opp strength avg', lambda r: r['opp_avg'], 'up'),
        ('Opponents', lambda r: r['opp_count'], 'up'),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Players'

    header_font = Font(bold=True)
    for c, (header, _accessor, _good) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    for r, row in enumerate(data['players'], start=2):
        for c, (_header, accessor, _good) in enumerate(cols, start=1):
            ws.cell(row=r, column=c, value=accessor(row))

    last_row = len(data['players']) + 1
    if last_row >= 2:
        for c, (_header, _accessor, good) in enumerate(cols, start=1):
            if good is None:
                continue
            letter = get_column_letter(c)
            ws.conditional_formatting.add(
                f'{letter}2:{letter}{last_row}', _color_scale(good),
            )

    # Freeze the header + first two identity columns; size columns to their headers.
    ws.freeze_panes = 'D2' if uses_teams else 'C2'
    for c, (header, _accessor, _good) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(9, min(len(header) + 2, 22))

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{subdomain or "tournament"}_stats.xlsx"'
    return response

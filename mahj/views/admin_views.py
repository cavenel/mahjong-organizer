import hashlib
import io
import logging
import random
import re
import time
import traceback
from collections import Counter, defaultdict, namedtuple
from datetime import datetime

import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from unidecode import unidecode

from django.conf import settings
from django.core.exceptions import ValidationError
from django.contrib.auth import logout
from django.db import IntegrityError, transaction
from django.db.models import F
from django.http import Http404, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, HttpResponseNotAllowed, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404
from django.template import loader
from django.utils.html import escape

from ..models import CARD_CSS_MAX, CeremonyState, Hand, Membership, Player, ScoreSheet, Seat, PublishedRound, Schedule, Screen, ScreenMode, Tenant
from ..signals import broadcast_display, broadcast_publish_state, invalidate_leaderboard
from .helpers import (
    BASE_DIR, VIEW_AS_SESSION_KEY, _TENANT_ROLES, acting_superuser, current_membership,
    get_counter, get_tenant, get_tournament, has_role, int_param, is_tenant_admin, json_body,
    method_not_allowed, real_is_tenant_admin, set_counter, tenant_admin_required,
    tenant_role_required,
)
from .user_admin import TENANT_ROLES, reauth_ok, tenant_admin_and_reauthed

logger = logging.getLogger(__name__)

def _display_redirect(open_panel=None):
    """Where a display-page action sends the operator back to.

    `open_panel` names a collapsible panel to reopen — passed as a query parameter,
    not a URL fragment. A fragment never reaches the server, so the panel could only
    be reopened by JS after the page had already rendered it shut; a parameter lets
    the template render it open in the first place, with no flicker and nothing to
    depend on. Adding or removing a screen lands the operator back on the screen grid
    they just changed rather than a collapsed panel.
    """
    url = 'admin?page=display'
    return f'{url}&open={open_panel}' if open_panel else url


def _reauth_gate(request, next_page=None):
    """The "confirm your password" panel shown in place of a sudo-gated page.

    A borrowed or unattended session must re-confirm before user management, tenant
    management or the database restore is even displayed. A link-only admin has no
    password to confirm, so the panel says so rather than offering a form they can't
    use. `next_page` is where to return to afterwards; the user console is the
    default landing page, so it passes none.
    """
    context = {"link_only": not request.user.has_usable_password()}
    if next_page:
        context["reauth_next"] = next_page
    return loader.get_template('mahj/admin_users_reauth.html').render(context, request)


def _sheet_state_keys(tenant, as_strings=False):
    """(validated, in-progress) score-sheet keys for this tenant.

    A sheet is *validated* when its ScoreSheet says so, and *in progress* when it has
    at least one played hand but is not validated — so the two sets never overlap and
    a table falls in at most one. Two callers want two key shapes: the welcome
    dashboard groups by round on `(round_nb, table_nb)` tuples, while the score grid
    matches the "<round>-<table>" strings its template builds. One query pair either
    way, and one definition of the rule.
    """
    def key(rn, tn):
        return f'{rn}-{tn}' if as_strings else (rn, tn)

    validated = {
        key(rn, tn)
        for rn, tn in ScoreSheet.objects.filter(tenant=tenant, validated=True)
                                        .values_list('round_nb', 'table_nb')
    }
    filled = {
        key(rn, tn)
        for rn, tn in Hand.objects.filter(tenant=tenant, win_by__isnull=False)
                                  .values_list('round_nb', 'table_nb').distinct()
    } - validated
    return validated, filled


# Human labels for the tenant-role flags, shown in the user-management console.
TENANT_ROLE_LABELS = {'scorer': 'Scorer', 'display_op': 'Display operator', 'publisher': 'Publisher'}
from ..scoring import WIND_LETTERS, _attach_players, _country_flag, pad_scores, rounds_played
from .scoring import (
    scores_per_player_rows,
    scores_per_table_grid,
)


# Friendly labels for the editable tournament tournament, matching the field
# labels on the display admin page, so a rejected save names the field the way
# the operator sees it ("On-screen message", not "welcome").
_TOURNAMENT_LABELS = {
    "welcome": "On-screen message",
    "title": "Title",
    "fullname": "Full tournament name",
    "city": "City",
    "period": "Period",
    "rules": "Rules",
    "home_country": "Home nation",
    "countrycourt": "Federation code",
    "card_format": "Card format",
    "card_theme": "Card theme",
    "card_css": "Card CSS",
}

# Which TournamentSettings fields each page's set_tournament may write. An
# explicit allowlist (not a denylist): a display operator tuning layout must not
# be able to reach structural fields (nb_rounds, rules, has_teams, …), and no
# page may write the server-authoritative `counter`. The sets mirror the fields
# each page's form actually submits (the `tournament-input` controls).
DISPLAY_SETTINGS_FIELDS = frozenset({
    "rotation_time", "score_lines", "total_columns", "welcome", "zoom",
})
TOURNAMENT_SETTINGS_FIELDS = frozenset({
    "city", "countrycourt", "fullname", "has_teams", "home_country", "is_test",
    "nb_rounds", "period", "rules", "title", "total_time",
    # Printed player cards, edited on the card-design page (which posts to this
    # same action, so the two pages share one handler and one allowlist).
    "card_format", "card_theme", "card_css",
})


def _screen_mode_or_404(tenant, raw_id):
    """Resolve a ScreenMode by id within ``tenant``, 404ing on a missing or
    non-numeric id (a stale button in a second tab, or a crafted request) rather
    than raising DoesNotExist/ValueError as a 500."""
    try:
        mode_id = int(raw_id)
    except (TypeError, ValueError):
        raise Http404("No such mode")
    return get_object_or_404(ScreenMode, tenant=tenant, id=mode_id)


def _name_is_round(name):
    """Guess whether an imported schedule row is a playing round from its name.

    Used only at import time to seed ``Schedule.is_round`` when the template
    omits the optional "Is round" column; staff can correct it afterwards in
    the settings UI.
    """
    name = name or ""
    return "round" in name.lower() or "session" in name.lower()


# Seating cells in the shipped template mirror the low half into the high half
# with formulas like "=14+8" (draw 22). openpyxl returns None for a formula cell
# unless the workbook carries Excel's cached result, and re-saving through
# openpyxl (as export does) drops that cache — so the seating sheet must be read
# formula-aware, evaluating the simple integer arithmetic here rather than
# treating an un-cached formula as an empty seat.
_SEAT_ARITH = re.compile(r'^\d+(\s*[+-]\s*\d+)*$')


def _seat_draw_number(raw):
    """The draw number seated in one seating cell. Accepts a literal number or a
    simple additive formula ("=14+8"); anything else (blank, or a formula we
    don't evaluate) is 0, the empty-seat sentinel, which the chart validation
    below then rejects for a full field."""
    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        return int(raw)
    s = str(raw).strip().lstrip('=')
    if not s:
        return 0
    if _SEAT_ARITH.match(s):
        # Only digits and +/- reached here, so summing the signed terms is safe.
        return sum(int(tok) for tok in re.findall(r'[+-]?\d+', s))
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return 0


def _pretty_view(view):
    """Human label for a stored screen view string. Mirrors the prettyView()
    used client-side on the display admin page, so a mode's saved views read the
    same whether rendered by Django or refreshed live by JS.
    Grammar: "black" | "welcome" | "counter" | "announcement" | "schedule"
    | "scores:<density>:<page>"."""
    if not view or view in ("black", "null"):
        return "Blank"
    if view == "welcome":
        return "Welcome"
    if view == "counter":
        return "Counter"
    if view == "announcement":
        return "Announcement"
    if view == "schedule":
        return "Schedule"
    if view.startswith("scores:"):
        parts = view.split(":")
        density = parts[1] if len(parts) > 1 and parts[1] in ("totals", "teams") else "detailed"
        page = parts[2] if len(parts) > 2 else "all"
        page_label = "all (rotating)" if page in ("", "all") else "page " + page
        return f"Standings — {density}, {page_label}"
    return view


def _mode_breakdowns(modes, screens):
    """Decorate each saved mode with the per-screen views it would apply, so the
    admin can show what clicking a mode does.

    `views` is a JSON list of view strings in screen order. A mode is a full-room
    snapshot: applying it sets every screen, with screens beyond the saved list
    (added after the mode was saved) going blank. is_active mirrors that — the
    mode's padded views must equal every screen's current view, so a mode reads
    as active exactly when re-clicking it would be a no-op. (Initial paint only;
    JS keeps the highlight current after live edits.)"""
    current = [str(s.view) or "black" for s in screens]
    # Positional label (/1, /2…) plus the operator's name when the screen was renamed.
    labels = [f"/{i + 1}" + (f" — {s.friendly_name}" if s.friendly_name else "")
              for i, s in enumerate(screens)]
    out = []
    for mode in modes:
        views = mode.views if isinstance(mode.views, list) else []
        normalised = [v or "black" for v in views]
        padded = [normalised[i] if i < len(normalised) else "black"
                  for i in range(len(current))]
        out.append({
            "id": mode.id,
            "name": mode.name,
            "rows": [{"label": labels[i], "pretty": _pretty_view(padded[i])}
                     for i in range(len(current))],
            "views_json": json.dumps(normalised, separators=(',', ':')),
            "is_active": bool(current) and padded == current,
        })
    return out


def publisher_overview_rows(tenant, tournament):
    """Per-round summary for the Publisher overview page.

    One dict per round 1..nb_rounds with the counts shown in the table and the
    underlying per-table id lists, so the page can keep the aggregates accurate
    as it applies the same scorer.* / publish.state WebSocket deltas the Scoring
    page consumes (rather than re-querying on every keystroke). The sheet state
    mirrors the Scoring page badge exactly: a validated sheet is never also
    counted as in-progress (filled_keys already excludes validated tables).
    """
    nb_rounds = tournament.nb_rounds or 0

    total_per = defaultdict(int)
    mp_per = defaultdict(int)
    for rn, tn, mp in Seat.objects.filter(tenant=tenant).values_list('round_nb', 'table_nb', 'minipoints'):
        total_per[(rn, tn)] += 1
        if mp is not None:
            mp_per[(rn, tn)] += 1

    tables_per_round = defaultdict(int)
    scored_tables = defaultdict(list)  # round -> [table_nb] (every seat has minipoints)
    for (rn, tn), total in total_per.items():
        tables_per_round[rn] += 1
        if total > 0 and mp_per[(rn, tn)] == total:
            scored_tables[rn].append(tn)

    validated_keys, filled_keys = _sheet_state_keys(tenant)

    validated_tables = defaultdict(list)
    for rn, tn in validated_keys:
        validated_tables[rn].append(tn)
    inprogress_tables = defaultdict(list)
    for rn, tn in filled_keys:
        inprogress_tables[rn].append(tn)

    published = set(
        PublishedRound.objects.filter(tenant=tenant).values_list('round_nb', flat=True)
    )

    rows = []
    for r in range(1, nb_rounds + 1):
        total = tables_per_round.get(r, 0)
        scored = sorted(scored_tables.get(r, []))
        rows.append({
            'round_nb': r,
            'tables_total': total,
            'scored_tables': scored,
            'inprogress_tables': sorted(inprogress_tables.get(r, [])),
            'validated_tables': sorted(validated_tables.get(r, [])),
            'published': r in published,
            'complete': total > 0 and len(scored) == total,
        })
    return rows


@tenant_admin_required
def admin_print_EMA(request):
    tournament = get_tournament(request)
    if tournament.rules == "MCR":
        wb = load_workbook(filename=str(BASE_DIR / "files/report_template.xlsx"), data_only=True)
        sheet_ranges = wb['MCR template']
    else:
        wb = load_workbook(filename=str(BASE_DIR / "files/report_template_Riichi.xlsx"), data_only=True)
        sheet_ranges = wb['Riichi template']
    standings = scores_per_player_rows(request, True)
    for row, player in enumerate(standings):
        items = [
            tournament.fullname,
            len(standings),
            player["pos"],
            player["first_name"],
            player["last_name"],
            player["EMA_ID"],
            player["total"]["tp"],
            player["total"]["mp"],
            "YES" if player["EMA_ID"] != "" else "NO",
            player["flag"].upper(),
            datetime.today().strftime('%d/%m/%Y'),
            tournament.countrycourt,   # Countrycourt: organising federation code (settings)
            tournament.city,
            2,
            tournament.title,
            tournament.rules,   # discipline column: "MCR" or "Riichi", not hardcoded
            tournament.period,
            2,
            "NO",
        ]
        for col, item in enumerate(items):
            sheet_ranges.cell(row=row + 2, column=col + 1).value = item
    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="EMA_report.xlsx"'
    return response


def _normalize_ema_id(raw):
    """The stored form of an EMA number: blank, or exactly eight digits.

    The id is optional — many competitors have none — so blank stays blank. A
    present-but-unparseable value is the organizer's mistake to surface, not to
    swallow: blanking it silently flips the EMA-report flag, and the EMA export
    reads the stored value back with ``int()``, so a free-text one crashes it.
    Zero-padded because that is the canonical form — an unpadded "1234" wouldn't
    match the same competitor imported from a template.

    Raises ``ValueError`` (via ``int``/``float``) for the caller to report in its
    own idiom: a template error on import, a 400 in the player editor.
    """
    if raw is None or not str(raw).strip():
        return ""
    return f"{int(float(str(raw).strip())):08d}"


def _ema_id_for_export(stored):
    """A stored EMA id as the template's number column wants it.

    Normally an int, since _normalize_ema_id keeps the field digits-only. A row
    predating that rule can still hold free text, so pass those through verbatim:
    dropping them would lose data out of what is meant to be a backup, and
    ``int()`` on them is what used to 500 the whole export. A re-import then
    reports the offending competitor by name.
    """
    if not stored:
        return None
    try:
        return int(stored)
    except (TypeError, ValueError):
        return stored


class TemplateImportError(Exception):
    """A validation problem in an uploaded tournament template that the organizer
    can fix (bad cell, wrong player count, missing seating sheet). Its message is
    shown to them verbatim; any other exception is treated as an unexpected error
    and reported with a traceback instead."""


def _player_row_has_name(last_name_raw, first_name_raw):
    """A Players row is a real competitor as long as it carries any name — a
    mononym (only a first name, or only a surname) counts, and a fully-blank
    row is a spacer to skip. One definition, used by the pre-checks and the
    import loop, so they can't disagree on what a competitor is."""
    return any(isinstance(v, str) and v.strip() for v in (last_name_raw, first_name_raw))


# The Options sheet's value cells, in order: fullname, title, nb_rounds, city,
# period, rules, published rounds, withheld rounds. The last two belong to the
# score half of the workbook and are blank in a setup-only export.
_OPTIONS_ROWS = 8


def _assign_short_names(players):
    """Set ``short_name`` on every Player in ``players`` (in memory, no DB write):
    the first name alone, or first name + the shortest surname prefix that
    separates competitors who share a first name ("Chris D.", growing to
    "Chris Dere." if two share "Der"). Disambiguation is across the whole list, so
    callers pass the tenant's full roster and persist with one bulk_update."""
    by_first = defaultdict(list)
    for player in players:
        by_first[unidecode(player.first_name).lower()].append(player)
    for group in by_first.values():
        if len(group) == 1:
            group[0].short_name = group[0].first_name
            continue
        for player in group:
            if not player.last_name:
                player.short_name = player.first_name
                continue
            n = 1
            while n < len(player.last_name):
                prefix = unidecode(player.last_name[:n]).lower()
                if sum(unidecode(p.last_name[:n]).lower() == prefix
                       for p in group) == 1:
                    break
                n += 1
            player.short_name = f"{player.first_name} {player.last_name[:n]}."


def _options_column(wb):
    """The Options sheet's value column, always ``_OPTIONS_ROWS`` long.

    The workbook is opened ``read_only=True``, so openpyxl yields only the rows that
    exist rather than padding to ``max_row`` — a sheet with four rows gave a
    four-element list. The pre-check guarded its own index and the main parse did not,
    so a short Options sheet passed the pre-check and then raised IndexError *after*
    the deletes, landing in the wipe-to-empty handler. Padding is the same answer a
    blank cell already gets.
    """
    vals = [row[1] for row in wb['Options'].iter_rows(
        min_row=1, max_row=_OPTIONS_ROWS, max_col=2, values_only=True)]
    return vals + [None] * (_OPTIONS_ROWS - len(vals))


def _precheck_template(wb):
    """The cheap checks that catch a wrong or old-format workbook, run BEFORE
    the import deletes anything: required sheets present, a readable rounds
    count, a playable player count, no colliding draw numbers. Each failure
    raises TemplateImportError, and the caller answers it with the tournament
    untouched — converting "wrong file erased my tournament" into "wrong file
    was rejected". A workbook that passes these and still fails deeper in the
    parse keeps the deliberate wipe-to-empty (import is a full replace): these
    checks exist to catch the wrong *file*, not every wrong cell."""
    for name in ('Options', 'Players', 'Schedule'):
        if name not in wb.sheetnames:
            raise TemplateImportError(
                f"The workbook has no '{name}' sheet — is this a tournament "
                f"template? Export the current tournament, or download the blank "
                f"template, to get a file in the expected format."
            )

    opt_vals = _options_column(wb)
    # A blank/zero rounds count would create no seating at all and "succeed"
    # with nothing playable. int() also normalises Excel's float (5.0 -> 5).
    try:
        nb_rounds = int(opt_vals[2])
    except (TypeError, ValueError):
        nb_rounds = 0
    if nb_rounds < 1:
        raise TemplateImportError(
            "The Options sheet must set the number of rounds to at least 1 "
            "(it was blank, zero or unreadable) — otherwise no seating chart "
            "is created."
        )

    player_rows = [row for row in wb['Players'].iter_rows(
                       min_row=2, max_col=6, values_only=True)
                   if _player_row_has_name(row[0], row[1])]
    if not player_rows:
        raise TemplateImportError(
            "The Players sheet lists no competitors. Fill in the player list "
            "(a row counts once it carries a name)."
        )
    if len(player_rows) % 4:
        raise TemplateImportError(
            f"The Players sheet lists {len(player_rows)} competitors; mahjong "
            f"seats four to a table, so the count must be a multiple of 4."
        )

    # Duplicate draw numbers would violate the per-tenant unique constraint
    # mid-load — a traceback and a wiped tournament. Blank cells are fine (a
    # player not yet drawn in); unparseable ones are named per-competitor by
    # the main parse.
    draws = []
    for row in player_rows:
        try:
            draws.append(int(row[5]))
        except (TypeError, ValueError):
            continue
    duplicated = sorted({d for d in draws if draws.count(d) > 1})
    if duplicated:
        raise TemplateImportError(
            f"Draw number {duplicated[0]} is assigned to more than one "
            f"competitor in the 'rand' column. Each draw number may appear "
            f"only once."
        )


@tenant_admin_required
def admin_upload_from_template(request):
    tenant = get_tenant(request)
    if request.method == 'POST':
        attached_file = request.FILES.get("myfile", None)
        if attached_file is None:
            return options(request)

        # Read straight from the upload, never via a path on disk: any shared
        # staging file is a cross-tenant race, since two concurrent imports
        # would fight over it and one tenant could load the other's workbook.
        # data_only=False so the seating sheet's mirror formulas ("=14+8")
        # are read as formulas and evaluated below; a file with no cached
        # results would otherwise read those seats as empty. Options/Players/
        # Schedule hold plain values, so this doesn't change how they read.
        #
        # Loading and pre-checking happen before anything is deleted: a file
        # that isn't a workbook, or fails a pre-check, is rejected with the
        # tournament untouched. Only a workbook that gets past this point is
        # allowed to replace the tournament — and from there on, a failure
        # deliberately wipes to empty (see the except below).
        try:
            wb = load_workbook(attached_file, data_only=False, read_only=True)
            _precheck_template(wb)
            # Scores are optional: a workbook exported without them, or the blank
            # template, has no Scores tab and imports as an empty tournament.
            # Parsed here, before anything is deleted, so an unreadable score cell
            # is rejected with the tournament untouched.
            scores = _parse_score_tabs(wb)
        except TemplateImportError as exc:
            return options(request, error=(
                "Import failed — nothing was changed.<br/>{0}".format(escape(str(exc)))))
        except Exception as exc:
            return options(request, error=(
                "Import failed — nothing was changed.<br/>The file could not be "
                "read as an Excel workbook (.xlsx):<br/><code>{0}</code>".format(
                    escape(str(exc)))))

        try:
            # One transaction for the whole wipe-and-load. The except below covers a
            # failure we can catch; this covers the one we can't — the worker being
            # killed or losing the database mid-import — which would otherwise commit a
            # genuinely half-imported tournament the organizer cannot tell from a good
            # one. The broadcasts stay outside it, so nothing is announced until the
            # import has actually committed.
            with transaction.atomic():
                Player.objects.filter(tenant=tenant).delete()

                Schedule.objects.filter(tenant=tenant).delete()
                sched_sheet = wb['Schedule']
                schedule_rows = list(sched_sheet.iter_rows(min_row=2, max_col=4, values_only=True))
                # Column 4 ("Is round") is optional; when it is missing (an older or
                # hand-edited template leaves the cell None) fall back to guessing from
                # the name (a row named "Round N" / "Session N" is a playing round).
                # Either way staff can correct any misclassification afterwards in
                # Tournament settings.
                Schedule.objects.bulk_create([
                    Schedule(tenant=tenant, day=row[0], time=row[1], name=row[2],
                             is_round=bool(row[3]) if row[3] is not None else _name_is_round(row[2]))
                    for row in schedule_rows if row[0] is not None
                ])

                opt_vals = _options_column(wb)
                tournament = get_tournament(request)
                tournament.fullname = opt_vals[0] or ""
                tournament.title = opt_vals[1] or ""
                # Parseability and the >= 1 floor were established by
                # _precheck_template before anything was deleted; int() also
                # normalises Excel's float (5.0 -> 5) before it reaches
                # iter_rows' max_row below.
                nb_rounds = int(opt_vals[2])
                tournament.nb_rounds = nb_rounds
                tournament.city = opt_vals[3] or ""
                tournament.period = opt_vals[4] or ""
                tournament.rules = opt_vals[5] or "MCR"
                tournament.save()

                # Player list: one Player per real person. The optional 'rand' column is a
                # pre-assigned draw number; when present the person is linked to their
                # seats below, otherwise the draw is made later (randomize / team draw).
                players_sheet = wb['Players']
                player_rows = list(players_sheet.iter_rows(min_row=2, max_col=6, values_only=True))
                player_objs = []
                any_team = False
                all_have_team = True
                for row in player_rows:
                    last_name_raw, first_name_raw, ema_raw, country, team_raw, rand_raw = row
                    # Skip fully-blank spacer / trailing rows and keep scanning
                    # (the shared rule also drives the pre-check's player count).
                    if not _player_row_has_name(last_name_raw, first_name_raw):
                        continue
                    # Keep the organizer's casing (strip only) — title-casing mangles
                    # "McDonald" and "van der Berg". The raw first/last are stored; the
                    # canonical display is "First Last".
                    first = first_name_raw.strip() if isinstance(first_name_raw, str) else ""
                    last = last_name_raw.strip() if isinstance(last_name_raw, str) else ""
                    full_name = f"{first} {last}".strip()
                    # Coerce to str (a numeric team cell would otherwise crash the whole
                    # import) and collapse internal whitespace, so "Team  A" and "Team A"
                    # aren't grouped as two different teams.
                    team = " ".join(str(team_raw).split()) if team_raw is not None else ""
                    if team:
                        any_team = True
                    else:
                        all_have_team = False
                    # The EMA id is optional (many competitors have none) -> blank
                    # silently. A present-but-unparseable value is the organizer's
                    # mistake to surface, not to swallow: blanking it also silently
                    # flips the EMA-report flag.
                    try:
                        ema = _normalize_ema_id(ema_raw)
                    except (TypeError, ValueError):
                        raise TemplateImportError(
                            f"Competitor '{full_name}' has an invalid EMA number "
                            f"'{ema_raw}'. Enter digits only, or leave the cell blank."
                        )
                    # The optional 'rand' column is the competitor's draw number. It is
                    # 1-based; 0 (or negative) collides with the empty-seat sentinel and
                    # would attach the player to every empty seat. The draw lives on the
                    # Player; seats are keyed by it. None until drawn.
                    if rand_raw is None:
                        draw_number = None
                    else:
                        try:
                            draw_number = int(rand_raw)
                        except (TypeError, ValueError):
                            raise TemplateImportError(
                                f"Competitor '{full_name}' has an invalid draw number "
                                f"'{rand_raw}'. Use a whole number from 1, or leave it blank."
                            )
                        if draw_number < 1:
                            raise TemplateImportError(
                                f"Competitor '{full_name}' has draw number {draw_number}; "
                                f"draw numbers start at 1."
                            )
                    player_objs.append(Player(
                        tenant=tenant, full_name=full_name, first_name=first,
                        last_name=last, EMA_ID=ema, country=country or "", team=team,
                        draw_number=draw_number,
                    ))
                if any_team and not all_have_team:
                    raise TemplateImportError(
                        "Some competitors have a team and some don't. When teams are "
                        "used every competitor must have one — fill in the blank team cells."
                    )
                # Teams are always groups of four. A team that comes out a different
                # size is a player-list mistake — most often a typo or case mismatch that
                # split one team in two ("Sweden" vs "sweden"), which the size check
                # catches without silently merging genuinely distinct names.
                if any_team:
                    sizes = Counter(p.team for p in player_objs)
                    wrong = sorted(name for name, n in sizes.items() if n != 4)
                    if wrong:
                        raise TemplateImportError(
                            f"Team '{wrong[0]}' has {sizes[wrong[0]]} competitor(s); every "
                            f"team must have exactly 4. Check for a typo or case mismatch "
                            f"in the team name."
                        )
                Player.objects.bulk_create(player_objs)

                # The player list is all-or-nothing on teams (enforced just above), so the
                # presence of any team name is the single signal for has_teams, which
                # gates team standings/columns/printouts everywhere.
                tournament.has_teams = any_team
                tournament.save()

                # bulk_create skips Player.save(), so set short_name here in one bulk_update.
                _assign_short_names(player_objs)
                Player.objects.bulk_update(player_objs, ['short_name'])

                Hand.objects.filter(tenant=tenant).delete()
                ScoreSheet.objects.filter(tenant=tenant).delete()
                Seat.objects.filter(tenant=tenant).delete()
                # Every score is about to be replaced by the workbook's, so any round
                # published for the previous tournament is stale — unpublish them all.
                # A workbook carrying scores re-publishes what its Options sheet
                # names (see _load_score_tabs).
                PublishedRound.objects.filter(tenant=tenant).delete()
                nb_players = len(player_objs)
                nb_tables = nb_players // 4
                # The seating sheet is optional: a workbook may carry only the player list,
                # schedule and settings, leaving the tournament without a chart until
                # one is built on the Seating page. When a "<N> players" sheet *is*
                # present it is read (and validated) here.
                # Consumed seat by seat as the chart is built; whatever is left
                # over named a seat the chart hasn't got, which is checked below.
                # Both stay defined when the workbook carries no seating sheet — an
                # uploaded score tab then finds no seat at all and is rejected.
                seat_scores = dict(scores['seats']) if scores else {}
                seats_to_create = []
                sheet_name = '{0} players'.format(nb_players)
                if sheet_name in wb.sheetnames:
                    seating_sheet = wb[sheet_name]
                    # Materialize the full seating sheet once (rows 3..3+nb_rounds-1,
                    # cols 2..2+5*nb_tables-1). Each cell holds the draw number seated.
                    seating_rows = list(seating_sheet.iter_rows(
                        min_row=3, max_row=2 + tournament.nb_rounds,
                        min_col=2, max_col=1 + 5 * nb_tables,
                        values_only=True,
                    ))
                    seats_to_create = []
                    expected = set(range(1, nb_players + 1))
                    for round_idx, row in enumerate(seating_rows):
                        round_draws = []
                        for table_nb in range(nb_tables):
                            for wind in range(4):
                                draw_number = _seat_draw_number(row[wind + 5 * table_nb])
                                round_draws.append(draw_number)
                                # The score the Scores tab recorded for this
                                # seat, if the workbook carries one. Applied as the
                                # Seat is built rather than updated afterwards:
                                # bulk_create is the only write, and an unscored
                                # seat is simply one the tab left blank.
                                recorded = seat_scores.pop(
                                    (round_idx + 1, table_nb + 1, wind + 1), {})
                                seats_to_create.append(Seat(
                                    tenant=tenant,
                                    draw_number=draw_number,
                                    round_nb=round_idx + 1,
                                    table_nb=table_nb + 1,
                                    wind=wind + 1,
                                    minipoints=recorded.get('minipoints'),
                                    tablepoints=recorded.get('tablepoints'),
                                    penalty=recorded.get('penalty', 0),
                                ))
                        # A valid chart seats every competitor 1..N exactly once each
                        # round. A mismatch means the sheet is the wrong size, has blank
                        # cells, or (the classic case) carries formulas with no cached
                        # result — better to reject it than load a half-empty chart.
                        if sorted(round_draws) != sorted(expected):
                            raise TemplateImportError(
                                f"The '{sheet_name}' seating sheet is not a valid chart: "
                                f"round {round_idx + 1} does not seat each competitor "
                                f"1–{nb_players} exactly once. Check the sheet for blank "
                                f"cells, duplicate numbers, or the wrong field size."
                            )
                    Seat.objects.bulk_create(seats_to_create, batch_size=500)

                # A score for a seat the chart doesn't have means the two halves of
                # the workbook disagree — a score tab kept from a tournament of a
                # different size, or a seating sheet that was edited out. Loading it
                # would silently drop results, so reject the file instead.
                if seat_scores:
                    round_nb, table_nb, wind = sorted(seat_scores)[0]
                    raise TemplateImportError(
                        f"The Scores sheet has a score for round {round_nb}, table "
                        f"{table_nb}, {WIND_LETTERS[wind - 1]}, which the seating "
                        f"chart does not seat. Check that the score tabs and the "
                        f"'{sheet_name}' sheet come from the same tournament."
                    )
                if scores:
                    _load_score_tabs(
                        tenant, scores, opt_vals,
                        {(s.round_nb, s.table_nb) for s in seats_to_create})
                wb.close()
        except Exception as exc:
            # Import is a full replace by design (scores included, whether the
            # workbook fills them in or leaves them empty), and a half-imported
            # tournament is worse than none — so any failure
            # leaves the tournament fully empty rather than half-loaded or silently
            # reverted to the old one. Wipe every player/seating/score table and
            # reset the settings the parse may have touched, so no half-branded
            # "ghost" tournament survives. The organizer fixes the file and retries.
            Player.objects.filter(tenant=tenant).delete()
            Hand.objects.filter(tenant=tenant).delete()
            ScoreSheet.objects.filter(tenant=tenant).delete()
            Seat.objects.filter(tenant=tenant).delete()
            PublishedRound.objects.filter(tenant=tenant).delete()
            Schedule.objects.filter(tenant=tenant).delete()
            tournament = get_tournament(request)
            tournament.fullname = ""
            tournament.title = ""
            tournament.nb_rounds = 0
            tournament.city = ""
            tournament.period = ""
            tournament.rules = "MCR"
            tournament.has_teams = False
            tournament.save()
            if isinstance(exc, TemplateImportError):
                # A validation problem we detected: show the actionable message.
                message = "Import failed — nothing was loaded.<br/>{0}".format(escape(str(exc)))
            else:
                # Unexpected error: keep the traceback to help diagnose it.
                message = "Error while creating tournament:<br/><code>{0}</code>".format(
                    escape(traceback.format_exc())
                )
            return options(request, error=message)

        # Outside the try on purpose. Both of these run *after* the transaction has
        # committed, so a failure in either says nothing about the import — but while
        # they sat inside the try, a Redis hiccup in invalidate_leaderboard landed in
        # the wipe handler above and emptied a tournament that had just imported
        # cleanly. The cache and the displays are best-effort; the data is not.
        try:
            invalidate_leaderboard(tenant.subdomain)
            broadcast_publish_state(tenant.subdomain, {'published_rounds': []})
        except Exception:
            logger.exception(
                "Import of %r succeeded but the cache/display refresh failed; the "
                "tournament is loaded and a later publish will resync the screens.",
                tenant.subdomain)

        return options(request)

    # GET (or any non-POST): show the settings page rather than falling through
    # to an implicit None return, which Django turns into a 500.
    return options(request)


def _export_filename(tournament, suffix=''):
    """Download name for an export: the tournament's short name reduced to a safe
    charset (the title is free text), falling back to "tournament"."""
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (tournament.title or ""))
    return (safe.strip("_") or "tournament") + suffix


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="{0}.xlsx"'.format(filename)
    return response


def _write_header(sheet, headers, freeze=None, header_row=1):
    """Write a bold, wrapped header row and size each column to its label."""
    for col, label in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        sheet.column_dimensions[get_column_letter(col)].width = max(9, min(len(label) + 2, 24))
    if freeze:
        sheet.freeze_panes = freeze


def scoring_has_started(tenant):
    """True once anything has been scored — a seat carrying minipoints, or any hand
    row. What gates the export's "include scores" checkbox, and what the import
    page's confirm dialog uses to name what an upload would replace."""
    return (Seat.objects.filter(tenant=tenant, minipoints__isnull=False).exists()
            or Hand.objects.filter(tenant=tenant).exists())


# What one played hand was: the three outcomes Hand.win_by encodes, as a label for
# the export's Result column. Informational — the importer reads the winner's wind
# (blank for a draw) and the discarder's, which is what the outcome is stored as.
def _hand_result(hand):
    if hand.is_draw:
        return 'Draw'
    return 'Self-draw' if hand.is_self_draw else 'Discard'


# How far a score sheet got, as the Scores tab's "Sheet" column spells it. Blank is
# a third state: no ScoreSheet row at all, i.e. a table nobody has opened.
_SHEET_VALIDATED = 'validated'
_SHEET_STARTED = 'started'

# Wind names the Scores / Score sheets tabs accept, on top of the 'E'/'S'/'W'/'N'
# the export writes and a plain 1-4: a hand-filled workbook is likely to spell them.
_WIND_NAMES = {'east': 1, 'south': 2, 'west': 3, 'north': 4}


def _tab_cell_error(tab, lineno, message):
    return TemplateImportError(
        "'{0}' sheet, row {1}: {2}".format(tab, lineno, message))


def _tab_rows(wb, tab, required):
    """Iterate a score tab as (row number, {column name: value}).

    Columns are matched by their header text, not by position, so a workbook whose
    rules leave out the TP column — or one an organizer has rearranged — still
    reads. Yields data rows only; the header itself is consumed here.
    """
    rows = wb[tab].iter_rows(values_only=True)
    header = next(rows, None) or ()
    index = {cell.strip().casefold(): i
             for i, cell in enumerate(header) if isinstance(cell, str)}
    missing = [name for name in required if name not in index]
    if missing:
        raise TemplateImportError(
            "The '{0}' sheet has no '{1}' column. Export the tournament again to "
            "get a file in the expected format.".format(tab, missing[0]))
    for lineno, row in enumerate(rows, start=2):
        yield lineno, {name: (row[i] if i < len(row) else None)
                       for name, i in index.items()}


def _tab_number(raw, tab, lineno, label, cast=int, blank_ok=False):
    """One numeric cell of a score tab. Blank is None when the column allows it."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        if blank_ok:
            return None
        raise _tab_cell_error(tab, lineno, "{0} is blank.".format(label))
    try:
        # via float so Excel's "4.0" reads as 4 for an integer column.
        return cast(float(raw)) if cast is int else cast(raw)
    except (TypeError, ValueError):
        raise _tab_cell_error(
            tab, lineno, "{0} is '{1}', which is not a number.".format(label, raw))


def _tab_wind(raw, tab, lineno, label, blank_ok=False):
    """One wind cell: E/S/W/N, East/South/West/North or 1-4, as 1-4."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        if blank_ok:
            return None
        raise _tab_cell_error(tab, lineno, "{0} is blank.".format(label))
    if isinstance(raw, str):
        text = raw.strip().casefold()
        if text in _WIND_NAMES:
            return _WIND_NAMES[text]
        letters = [w.casefold() for w in WIND_LETTERS]
        if text in letters:
            return letters.index(text) + 1
    wind = _tab_number(raw, tab, lineno, label)
    if not 1 <= wind <= 4:
        raise _tab_cell_error(
            tab, lineno,
            "{0} is '{1}'; use E, S, W or N.".format(label, raw))
    return wind


def _parse_round_list(raw, label):
    """A comma-separated round list from the Options sheet, e.g. "1,2,3"."""
    if raw is None:
        return []
    rounds = []
    for token in str(raw).replace(';', ',').split(','):
        token = token.strip()
        if not token:
            continue
        try:
            rounds.append(int(float(token)))
        except (TypeError, ValueError):
            raise TemplateImportError(
                "The Options sheet's '{0}' value is '{1}'; list round numbers "
                "separated by commas, or leave it blank.".format(label, raw))
    return rounds


def _load_score_tabs(tenant, scores, opt_vals, chart_cells):
    """Create the ScoreSheet, Hand and PublishedRound rows an imported workbook's
    score tabs describe, for a tournament whose chart has just been loaded.

    The per-seat numbers are not written here — they are applied as each Seat is
    built (see admin_upload_from_template), so the chart stays one bulk_create.
    ``chart_cells`` is the set of (round, table) pairs the chart actually has: a
    hand or a sheet state for any other table means the workbook's halves disagree,
    the same mismatch a stray seat score is rejected for.
    """
    for cell in sorted(set(scores['sheets']) | {(h['round_nb'], h['table_nb'])
                                                for h in scores['hands']}):
        if cell not in chart_cells:
            raise TemplateImportError(
                "The score tabs carry results for round {0}, table {1}, which the "
                "seating chart does not have. Check that they come from the same "
                "tournament.".format(*cell))

    ScoreSheet.objects.bulk_create([
        ScoreSheet(tenant=tenant, round_nb=round_nb, table_nb=table_nb,
                   validated=state == _SHEET_VALIDATED)
        for (round_nb, table_nb), state in sorted(scores['sheets'].items())
    ])
    Hand.objects.bulk_create([
        Hand(tenant=tenant, **hand) for hand in scores['hands']
    ], batch_size=500)

    # Publication state travels with the scores: a workbook holding a finished
    # tournament restores as a published one, so the public site shows what it
    # showed before rather than needing every round re-published by hand.
    withheld = set(_parse_round_list(opt_vals[7], 'Withheld rounds'))
    PublishedRound.objects.bulk_create([
        PublishedRound(tenant=tenant, round_nb=round_nb, withheld=round_nb in withheld)
        for round_nb in sorted(set(_parse_round_list(opt_vals[6], 'Published rounds')))
    ])


def _parse_score_tabs(wb):
    """Read the Scores / Score sheets tabs of an uploaded workbook, or None when it
    carries no scores (a setup-only template).

    Called before the import deletes anything, so an unreadable cell is rejected
    with the tournament untouched. Returns plain dicts keyed the way the seating
    chart is — (round, table, wind) and (round, table) — for the load to apply;
    whether those seats actually exist is only knowable once the chart is built,
    and is checked there. The Standings tab is derived and deliberately ignored.
    """
    if 'Scores' not in wb.sheetnames:
        return None

    seat_scores = {}
    sheet_states = {}
    for lineno, row in _tab_rows(wb, 'Scores', ('round', 'table', 'wind', 'mp')):
        if all(row.get(c) is None for c in ('round', 'table', 'wind')):
            continue    # blank spacer / trailing row
        key = (
            _tab_number(row['round'], 'Scores', lineno, 'Round'),
            _tab_number(row['table'], 'Scores', lineno, 'Table'),
            _tab_wind(row['wind'], 'Scores', lineno, 'Wind'),
        )
        if key in seat_scores:
            raise _tab_cell_error(
                'Scores', lineno,
                "round {0} table {1} {2} is listed twice.".format(
                    key[0], key[1], WIND_LETTERS[key[2] - 1]))
        # A blank MP is an unscored seat — the state every seat starts in, and what
        # a round the chart seats but nobody has played yet looks like.
        seat_scores[key] = {
            'minipoints': _tab_number(row['mp'], 'Scores', lineno, 'MP', blank_ok=True),
            'tablepoints': _tab_number(row.get('tp'), 'Scores', lineno, 'TP',
                                       cast=float, blank_ok=True),
            'penalty': _tab_number(row.get('penalty'), 'Scores', lineno, 'Penalty',
                                   blank_ok=True) or 0,
        }
        state = row.get('sheet')
        state = str(state).strip().casefold() if state is not None else ''
        if state in (_SHEET_VALIDATED, _SHEET_STARTED):
            sheet_states[key[:2]] = state
        elif state:
            raise _tab_cell_error(
                'Scores', lineno,
                "Sheet is '{0}'; use '{1}', '{2}' or leave it blank.".format(
                    row.get('sheet'), _SHEET_VALIDATED, _SHEET_STARTED))

    hands = []
    if 'Score sheets' in wb.sheetnames:
        seen = set()
        for lineno, row in _tab_rows(wb, 'Score sheets',
                                     ('round', 'table', 'hand', 'value')):
            if all(row.get(c) is None for c in ('round', 'table', 'hand')):
                continue
            cell = (
                _tab_number(row['round'], 'Score sheets', lineno, 'Round'),
                _tab_number(row['table'], 'Score sheets', lineno, 'Table'),
                _tab_number(row['hand'], 'Score sheets', lineno, 'Hand'),
            )
            if cell in seen:
                raise _tab_cell_error(
                    'Score sheets', lineno,
                    "round {0} table {1} hand {2} is listed twice.".format(*cell))
            seen.add(cell)
            # No winner's wind is how a draw is written (the Result column says so
            # too, but the wind is what the outcome is stored as). A draw has no
            # discarder, whatever the row claims.
            win_by = _tab_wind(row.get('winner wind'), 'Score sheets', lineno,
                               'Winner wind', blank_ok=True)
            win_from = _tab_wind(row.get('dealt in wind'), 'Score sheets', lineno,
                                 'Dealt in wind', blank_ok=True)
            hands.append({
                'round_nb': cell[0], 'table_nb': cell[1], 'hand_nb': cell[2],
                'points': _tab_number(row['value'], 'Score sheets', lineno, 'Value',
                                      blank_ok=True) or 0,
                'win_by': win_by or 0,
                'win_from': win_from if win_by else None,
            })

    return {'seats': seat_scores, 'sheets': sheet_states, 'hands': hands}


@tenant_admin_required
def admin_export_to_template(request):
    """Export the whole tournament as one Excel workbook, in the format
    admin_upload_from_template reads back.

    The workbook is built from scratch (not from MahjongTemplate.xlsx). Its setup
    sheets are Options, Players, Schedule and a single "<N> players" seating sheet
    for this tournament's field size; with ``?scores=1`` it also carries Scores
    (what each seat recorded), Score sheets (every played hand) and Standings (the
    ranked table). Everything the importer restores is written, so one file is a
    full round-trip: back the tournament up, edit it offline, upload it again.

    Standings is the exception — it is derived from the other two and the importer
    ignores it. It is written because the ranked table is the sheet an organizer
    actually reads, and its own first row says edits there do nothing.
    """
    tenant = get_tenant(request)
    tournament = get_tournament(request)
    # The checkbox beside the download. Nothing to include before play starts, so
    # an empty tournament exports its setup whatever the box says.
    include_scores = request.GET.get('scores') == '1' and scoring_has_started(tenant)

    wb = Workbook()

    # Options: label in col A (for humans), value in col B (what the importer reads).
    # The publication rows are part of the score half — a setup-only export leaves
    # them blank, which reads back as "nothing published", matching its empty sheets.
    published = withheld = ''
    if include_scores:
        rounds = list(PublishedRound.objects.filter(tenant=tenant).order_by('round_nb'))
        published = ','.join(str(r.round_nb) for r in rounds)
        withheld = ','.join(str(r.round_nb) for r in rounds if r.withheld)
    opt_sheet = wb.active
    opt_sheet.title = 'Options'
    for row, (label, value) in enumerate([
        ('Competition name', tournament.fullname),
        ('Short name (initials)', tournament.title),
        ('Number of rounds', tournament.nb_rounds),
        ('City', tournament.city),
        ('Period', tournament.period),
        ('Rules', tournament.rules),
        ('Published rounds', published),
        ('Withheld rounds', withheld),
    ], start=1):
        opt_sheet.cell(row=row, column=1, value=label)
        opt_sheet.cell(row=row, column=2, value=value)

    # Players: id order is the original player-list row order (matches the draw exports).
    # Columns 1-6 mirror the shipped template.
    players = list(Player.objects.filter(tenant=tenant).order_by('id'))
    players_sheet = wb.create_sheet('Players')
    players_sheet.append([
        'Last name', 'First name', 'EMA number', 'Country',
        'Team name (optional)', 'Random position (1 - # of players)',
    ])
    for player in players:
        players_sheet.append([
            player.last_name,
            player.first_name,
            # EMA_ID is stored zero-padded ("00001234"); export the number so
            # _normalize_ema_id reproduces it on re-import (a blank stays blank).
            _ema_id_for_export(player.EMA_ID),
            player.country,
            player.team or None,
            player.draw_number,
        ])

    # Schedule: Date / Time / Name; the "Is round" flag is appended so it round-trips
    # (the importer otherwise re-guesses it from the name).
    sched_sheet = wb.create_sheet('Schedule')
    sched_sheet.append(['Date', 'Time', 'Name', 'Is round'])
    for item in Schedule.objects.filter(tenant=tenant).order_by('id'):
        sched_sheet.append([item.day, item.time, item.name, item.is_round])

    # Seating: a single "<N> players" sheet for this field size, built from the real
    # seating. Draw numbers are laid out as round rows x table blocks of E/S/W/N,
    # with one spacer column between tables (col = 2 + wind + 5*table, both 0-based) —
    # exactly the layout admin_upload_from_template reads back.
    nb_players = len(players)
    nb_tables = nb_players // 4
    seats = list(Seat.objects.filter(tenant=tenant))
    if seats and nb_tables:
        seating_sheet = wb.create_sheet('{0} players'.format(nb_players))
        for table_nb in range(1, nb_tables + 1):
            base = 2 + 5 * (table_nb - 1)
            seating_sheet.cell(row=1, column=base, value='Table {0}'.format(table_nb))
            for wind, label in enumerate(['East', 'South', 'West', 'North']):
                seating_sheet.cell(row=2, column=base + wind, value=label)
        nb_rounds = max(tournament.nb_rounds, max(s.round_nb for s in seats))
        for round_nb in range(1, nb_rounds + 1):
            seating_sheet.cell(row=round_nb + 2, column=1, value='R{0}'.format(round_nb))
        for seat in seats:
            col = 2 + (seat.wind - 1) + 5 * (seat.table_nb - 1)
            seating_sheet.cell(row=seat.round_nb + 2, column=col, value=seat.draw_number)

    if include_scores:
        _write_score_sheets(wb, request, tenant, tournament, seats)

    suffix = '_scores' if include_scores else ''
    return _xlsx_response(wb, _export_filename(tournament, suffix))


def _write_score_sheets(wb, request, tenant, tournament, seats):
    """Add the score half of an export: Scores, Score sheets and Standings.

    Nothing is re-derived on the way out — the first two sheets are the Seat and
    Hand rows as stored, which is what makes them importable; Standings is the
    scoring page's ranked table, for reading only.
    """
    is_mcr = tournament.rules == 'MCR'
    seats = _attach_players(tenant, seats)
    # Which competitor sat where, and how far each sheet got: both are keyed on the
    # (round, table) pair the importer matches on.
    sheet_state = {}
    for sheet in ScoreSheet.objects.filter(tenant=tenant):
        sheet_state[(sheet.round_nb, sheet.table_nb)] = (
            _SHEET_VALIDATED if sheet.validated else _SHEET_STARTED)

    # --- Scores: the number recorded on every seat -------------------------------
    scores_sheet = wb.create_sheet('Scores')
    headers = ['Round', 'Table', 'Wind', 'Draw #', 'Player', 'MP', 'Penalty']
    if is_mcr:
        headers.append('TP')
    headers.append('Sheet')
    _write_header(scores_sheet, headers, freeze='D2')
    for seat in sorted(seats, key=lambda s: (s.round_nb, s.table_nb, s.wind)):
        values = [
            seat.round_nb, seat.table_nb,
            WIND_LETTERS[seat.wind - 1] if 1 <= seat.wind <= 4 else '',
            seat.draw_number, seat.player_name(),
            seat.minipoints, seat.penalty,
        ]
        if is_mcr:
            values.append(seat.tablepoints)
        values.append(sheet_state.get((seat.round_nb, seat.table_nb), ''))
        scores_sheet.append(values)

    # --- Score sheets: one row per played hand -----------------------------------
    # Seat winds resolve to names in memory: a Hand only stores the wind that won
    # and the wind that dealt in, and Hand.win_by_player() would be a query each.
    seat_by_cell = {(s.round_nb, s.table_nb, s.wind): s for s in seats}

    def _seat_name(round_nb, table_nb, wind):
        seat = seat_by_cell.get((round_nb, table_nb, wind))
        return seat.player_name() if seat else ''

    hands_sheet = wb.create_sheet('Score sheets')
    _write_header(hands_sheet, [
        'Round', 'Table', 'Hand', 'Result', 'Value',
        'Winner wind', 'Winner', 'Dealt in wind', 'Dealt in by',
    ], freeze='D2')
    hands = Hand.objects.filter(tenant=tenant).order_by('round_nb', 'table_nb', 'hand_nb')
    for hand in hands:
        if hand.win_by is None:
            continue    # unplayed placeholder row on a sheet still being entered
        winner = '' if hand.is_draw else WIND_LETTERS[hand.win_by - 1]
        dealt_in = WIND_LETTERS[hand.win_from - 1] if hand.win_from else ''
        hands_sheet.append([
            hand.round_nb, hand.table_nb, hand.hand_nb, _hand_result(hand), hand.points,
            winner,
            '' if hand.is_draw else _seat_name(hand.round_nb, hand.table_nb, hand.win_by),
            dealt_in,
            _seat_name(hand.round_nb, hand.table_nb, hand.win_from) if hand.win_from else '',
        ])

    # --- Standings: the ranked table, for reading ---------------------------------
    standings = scores_per_player_rows(request, True)
    # A chart can run longer than the configured round count (an imported chart, a
    # round added mid-tournament), so take whichever is further along.
    nb_rounds = max(tournament.nb_rounds or 0, rounds_played(standings))
    uses_teams = tournament.has_teams

    st_sheet = wb.create_sheet('Standings')
    st_sheet.cell(row=1, column=1, value=(
        'Computed from the Scores and Score sheets tabs — importing this workbook '
        'ignores what is written here.'))
    st_sheet.cell(row=1, column=1).font = Font(italic=True)
    headers = ['Rank', 'Player', 'EMA number', 'Country']
    if uses_teams:
        headers.append('Team')
    if is_mcr:
        headers.append('Total TP')
    headers.append('Total MP')
    for round_nb in range(1, nb_rounds + 1):
        headers.append('R{0} table'.format(round_nb))
        headers.append('R{0} MP'.format(round_nb))
        if is_mcr:
            headers.append('R{0} TP'.format(round_nb))
    _write_header(st_sheet, headers, freeze='C3', header_row=2)

    # Which table each competitor sat at in each round, so a score can be read
    # beside the table it was played at — keyed on (player, round), never on
    # list position (a chart needn't seat everyone every round).
    player_table = {(s.player_id, s.round_nb): s.table_nb for s in seats if s.player_id}
    for row in standings:
        values = [row['pos'], row['name'], row['EMA_ID'], row['country']]
        if uses_teams:
            values.append(row['team'])
        if is_mcr:
            values.append(row['total']['tp'])
        values.append(row['total']['mp'])
        for score in pad_scores(row['scores'], nb_rounds):
            # mp None is a round the competitor didn't play (pad_scores' empty
            # cell) — leave the table blank there too rather than naming a seat
            # they never took.
            played = score['mp'] is not None
            values.append(player_table.get((row['player_id'], score['round_nb'])) if played else None)
            values.append(score['mp'])
            if is_mcr:
                values.append(score['tp'])
        st_sheet.append(values)


@tenant_admin_required
def admin_generate_seating(request):
    """Build the seating chart in-app for the current player list, instead of reading
    it from an Excel sheet — so a field size the template doesn't cover still gets
    a chart. Replaces the seating chart (and clears scores, which are keyed by
    seat) but keeps the player list, draw and schedule: the chart is independent of who
    sits where. Returns the quality measures as JSON for the page to display."""
    if request.method != 'POST':
        return method_not_allowed()

    tenant = get_tenant(request)
    tournament = get_tournament(request)
    nb_players = Player.objects.filter(tenant=tenant).count()
    nb_rounds = tournament.nb_rounds or 0

    # Body carries the chosen method, a variation seed, and whether to apply (vs
    # just preview the measures). Absent body -> auto method, apply immediately.
    body = json_body(request)
    method = body.get('method', 'auto')
    try:
        seed = int(body.get('seed', 0))
    except (TypeError, ValueError):
        seed = 0
    try:
        # Number of best-effort attempts to search and keep the best of. A preview
        # searches many (default); applying a previewed chart passes tries=1 with
        # the winning seed to reproduce it exactly. Bounded so one request can't run
        # unboundedly (a wall-clock budget also caps it).
        tries = min(5000, max(1, int(body.get('tries', 400))))
    except (TypeError, ValueError):
        tries = 400
    do_apply = bool(body.get('apply', True))

    from .. import seating
    try:
        rows, meta = seating.generate(
            nb_players, nb_rounds, has_teams=tournament.has_teams,
            seed=seed, method=method, tries=tries)
    except seating.SeatingInfeasible as exc:
        return JsonResponse({'error': str(exc)}, status=400)

    m = seating.measure(rows, nb_players, nb_rounds, has_teams=tournament.has_teams)
    m['engine'] = meta['engine']
    payload = {'ok': True, 'applied': do_apply,
               'headline': seating.headline(m), 'measures': m,
               # The winning seed (best-effort) lets the page reproduce this exact
               # chart on apply, and reveals it to the organizer.
               'seed': meta.get('seed'), 'tries_run': meta.get('tries_run')}
    if not do_apply:
        # Preview only — the chart is deterministic for (method, seed), so applying
        # the same choice later reproduces exactly what was previewed.
        return JsonResponse(payload)

    with transaction.atomic():
        # Seating changed, so any entered scores/published rounds are stale.
        Hand.objects.filter(tenant=tenant).delete()
        ScoreSheet.objects.filter(tenant=tenant).delete()
        Seat.objects.filter(tenant=tenant).delete()
        PublishedRound.objects.filter(tenant=tenant).delete()
        Seat.objects.bulk_create([
            Seat(tenant=tenant, draw_number=draw_number, round_nb=round_nb,
                 table_nb=table_nb, wind=wind, minipoints=None, tablepoints=None)
            for (round_nb, table_nb, wind, draw_number) in rows
        ], batch_size=500)

    invalidate_leaderboard(tenant.subdomain)
    broadcast_publish_state(tenant.subdomain, {'published_rounds': []})
    return JsonResponse(payload)


@tenant_admin_required
def admin_team_draw(request):
    tenant = get_tenant(request)

    # The draw hands out seat numbers, so it needs a seating chart to draw for.
    # Without one every drawn number would be rejected at save time (after the
    # whole ceremony), so stop up front with a clear message instead.
    if not Seat.objects.filter(tenant=tenant).exists():
        template = loader.get_template('mahj/admin_seating_required.html')
        return HttpResponse(template.render(
            {'draw_title': 'Team Draw', 'draw_kind': 'team draw'}, request))

    players = list(Player.objects.filter(tenant=tenant).order_by('full_name'))

    # id order is creation order: the row order of an imported "Players" sheet, or
    # the order rows were added in the editor. The CSV export sorts on this so the
    # drawn numbers line up with the template rows.
    order = {p.id: i for i, p in enumerate(sorted(players, key=lambda p: p.id), start=1)}

    teams_dict = {}
    for p in players:
        if p.team:
            teams_dict.setdefault(p.team, []).append({
                "id": p.id,
                "original_index": order[p.id],
                "full_name": p.full_name,
                "short_name": p.short_name,
                "country": p.country,
                "flag": _country_flag(p.country),
                "EMA_ID": p.EMA_ID,
            })

    teams_list = [
        {"name": name, "players": players}
        for name, players in sorted(teams_dict.items())
    ]

    nb_teams = len(teams_list)

    # Existing draw, if any: the competitors who already have a draw number,
    # grouped by team.
    drawn = [p for p in players if p.draw_number is not None and p.team]

    saved_draw = []
    if drawn:
        draw_teams = {}
        for p in drawn:
            draw_teams.setdefault(p.team, []).append({
                "full_name": p.full_name,
                "rand_id": p.draw_number,
                "original_index": order.get(p.id, 0),
            })
        slot = 1
        for team_name in sorted(draw_teams.keys()):
            members = sorted(draw_teams[team_name], key=lambda x: x["rand_id"])
            saved_draw.append({
                "slot": slot,
                "team_name": team_name,
                "players": members,
            })
            slot += 1

    template = loader.get_template('mahj/admin_team_draw.html')
    # The page reads these through |json_script, which escapes the markup
    # characters json.dumps leaves alone — team and player names are
    # operator-entered, so a name containing `</script>` would inject script.
    context = {
        "teams": teams_list,
        "nb_teams": nb_teams,
        "saved_draw": saved_draw,
    }
    return HttpResponse(template.render(context, request))


@tenant_admin_required
def admin_team_draw_save(request):
    """Save a completed team draw: give each competitor the number they drew.

    The whole draw is replaced at once (the operator runs it as one ceremony), so
    the batch is validated before anything is written: every drawn number must be
    a real seat and no competitor or number may appear twice. Only then does the
    reassignment run, inside a single transaction, so a mid-write failure rolls
    back instead of leaving the draw half-wiped."""
    if request.method != 'POST':
        return method_not_allowed()

    tenant = get_tenant(request)
    assignments = json_body(request).get('assignments')  # [{player_id, rand_id}]
    if not isinstance(assignments, list) or not all(isinstance(a, dict) for a in assignments):
        return HttpResponse('Malformed request', status=400)
    # Coerced before anything compares them. A null or non-numeric drawn number
    # reached `sorted(set(draw_numbers) - valid)` below and raised TypeError
    # ordering None against int; a numeric *string* id passed the ORM lookup and
    # then missed the int-keyed `players` dict, a KeyError. Both 500s, on the one
    # request that writes the whole draw.
    try:
        player_ids = [int(a['player_id']) for a in assignments]
        draw_numbers = [int(a['rand_id']) for a in assignments]  # the drawn number
    except (KeyError, TypeError, ValueError):
        return HttpResponse('Malformed request', status=400)

    if len(set(player_ids)) != len(player_ids) or len(set(draw_numbers)) != len(draw_numbers):
        return HttpResponse('Each competitor and each draw number must appear once', status=400)

    valid = set(Seat.objects.filter(tenant=tenant).values_list('draw_number', flat=True))
    unknown = sorted(set(draw_numbers) - valid)
    if unknown:
        return HttpResponse('#{0} is not a valid seat'.format(unknown[0]), status=400)

    players = {p.id: p for p in Player.objects.filter(tenant=tenant, id__in=player_ids)}
    if len(players) != len(player_ids):
        return HttpResponse('Player list changed — reload and draw again', status=400)

    # Free the target draw numbers from any current holder and clear these
    # competitors' current numbers, then assign each their drawn number. Clearing
    # first keeps the per-tenant unique draw_number constraint satisfied.
    with transaction.atomic():
        Player.objects.filter(tenant=tenant, draw_number__in=draw_numbers).update(draw_number=None)
        Player.objects.filter(tenant=tenant, id__in=player_ids).update(draw_number=None)
        for player_id, draw_number in zip(player_ids, draw_numbers):
            players[player_id].draw_number = draw_number
        Player.objects.bulk_update(players.values(), ['draw_number'])

    # The draw decides who sits in which seat, so it changes what the standings,
    # the seating grid and the desktop HTML all say. views/scoring names the draw
    # in its invalidation contract; without this the projector serves the old
    # names for up to SUB_CACHE_TTL while the desktop page (which passes its own
    # seats) corrects immediately — two screens in one room disagreeing.
    invalidate_leaderboard(tenant.subdomain)
    return HttpResponse('OK')


@tenant_admin_required
def admin_player_draw(request):
    """Live individual draw: competitors arrive one at a time, physically draw a
    number, and the operator types it in. Each assignment is saved immediately
    (via admin_player_draw_assign) so the seating chart is live as the desk runs.
    The draw is recorded as Player.draw_number, same as Randomize / Team draw."""
    tenant = get_tenant(request)

    # No seating chart means no seat numbers to hand out, and every assignment
    # would be rejected by admin_player_draw_assign. Stop up front instead.
    if not Seat.objects.filter(tenant=tenant).exists():
        template = loader.get_template('mahj/admin_seating_required.html')
        return HttpResponse(template.render(
            {'draw_title': 'Player Draw', 'draw_kind': 'individual draw'}, request))

    all_players = list(Player.objects.filter(tenant=tenant).order_by('full_name'))

    # id order is creation order (imported sheet rows, or editor adds), so a CSV
    # export lines up with the template rows (matches the team-draw export).
    order = {p.id: i for i, p in enumerate(sorted(all_players, key=lambda p: p.id), start=1)}

    players = [{
        "id": p.id,
        "original_index": order[p.id],
        "full_name": p.full_name,
        "short_name": p.short_name,
        "country": p.country,
        "flag": _country_flag(p.country),
        "draw_number": p.draw_number,
    } for p in all_players]

    draw_numbers = sorted({s.draw_number for s in Seat.objects.filter(tenant=tenant)})

    template = loader.get_template('mahj/admin_player_draw.html')
    # |json_script in the template, not json.dumps here — see admin_team_draw.
    context = {
        "players": players,
        "draw_numbers": draw_numbers,
    }
    return HttpResponse(template.render(context, request))


def _draw_number_taken(draw_number, holder_name):
    """The 409 the live-draw page expects when a number is already gone: it reverts
    the row and reports who holds it."""
    taken_by = holder_name or 'someone else'
    return JsonResponse(
        {'ok': False,
         'error': f'#{draw_number} is already taken by {taken_by}',
         'holder': holder_name},
        status=409)


@tenant_admin_required
def admin_player_draw_assign(request):
    """Assign (or clear) one competitor's draw number for the live individual draw.

    POST {player_id, draw_number} where draw_number is the number the competitor
    physically drew, or null to clear it (undo). The availability check is done
    here under a row lock so two registration desks can't hand out the same
    number: the request fails if the number isn't a real draw slot or is already
    held by someone else."""
    if request.method != 'POST':
        return method_not_allowed()

    tenant = get_tenant(request)
    data = json_body(request)
    # Coerced up front rather than handed to the ORM raw: a non-numeric id used to
    # raise ValueError out of .get() below, and a non-scalar draw number a
    # TypeError out of the `in valid` test, both as 500s. The page reads
    # `error` off the JSON body, which is the shape FieldError renders.
    player_id = int_param(data, 'player_id')
    draw_number = int_param(data, 'draw_number', default=None)  # int to assign, None to clear

    with transaction.atomic():
        try:
            player = Player.objects.select_for_update().get(tenant=tenant, id=player_id)
        except Player.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Unknown competitor'}, status=404)

        if draw_number is None:
            player.draw_number = None
            player.save(update_fields=['draw_number'])
            invalidate_leaderboard(tenant.subdomain)  # the draw changed — see admin_team_draw_save
            return JsonResponse({'ok': True, 'player_id': player.id, 'draw_number': None})

        valid = set(Seat.objects.filter(tenant=tenant).values_list('draw_number', flat=True))
        if draw_number not in valid:
            return JsonResponse(
                {'ok': False, 'error': f'#{draw_number} is not a valid draw number'}, status=400)

        holder = (Player.objects
                  .select_for_update()
                  .filter(tenant=tenant, draw_number=draw_number)
                  .exclude(id=player.id)
                  .first())
        if holder is not None:
            return _draw_number_taken(draw_number, holder.full_name)

        player.draw_number = draw_number
        try:
            player.save(update_fields=['draw_number'])
        except IntegrityError:
            # When the number has no holder yet, select_for_update above locks no
            # row — there is nothing to lock — so two desks assigning the same
            # number at once both get here and the per-tenant unique constraint
            # rejects the loser. That's the same outcome as losing the check above,
            # so give the same 409 the client already knows how to handle rather
            # than a 500. Re-read to name whoever won.
            winner = (Player.objects
                      .filter(tenant=tenant, draw_number=draw_number)
                      .exclude(id=player.id)
                      .first())
            return _draw_number_taken(draw_number, winner.full_name if winner else '')
        invalidate_leaderboard(tenant.subdomain)  # the draw changed — see admin_team_draw_save
        return JsonResponse({'ok': True, 'player_id': player.id, 'draw_number': draw_number})


# The metadata fields the Player editor may change. draw_number is deliberately
# NOT here: it's the draw itself (set by Randomize / Team draw / import) and the
# seating chart is keyed by it, so reassigning it belongs to those tools, not to
# a free-text metadata edit.
_PLAYER_EDITABLE_FIELDS = ['first_name', 'last_name', 'EMA_ID', 'country', 'team']


@tenant_admin_required
def player_editor_save(request):
    """Persist inline edits from the Player editor table. Accepts JSON
    ``{"players": [{"id", <field>...}]}`` and bulk-updates the editable
    metadata; unknown ids are ignored, over-long values are rejected up front."""
    if request.method != 'POST':
        return method_not_allowed()

    tenant = get_tenant(request)
    rows = json_body(request).get('players', [])
    if not isinstance(rows, list):
        return HttpResponse('Malformed request body', status=400)
    rows = [r for r in rows if isinstance(r, dict)]
    # Coerced before the query: a non-numeric id raised ValueError out of the
    # `id__in` lookup as a 500, and a numeric *string* id survived the lookup only
    # to miss the int-keyed `by_id` below — the row was then skipped as unknown
    # and the editor reported a save that never happened. An id we can't read
    # names no competitor, so the batch is refused rather than part-applied.
    try:
        ids = [int(r.get('id')) for r in rows]
    except (TypeError, ValueError):
        return HttpResponse('Malformed request: every row needs a numeric id.', status=400)
    by_id = {p.id: p for p in Player.objects.filter(tenant=tenant, id__in=ids)}

    to_update = []
    names_changed = False
    for r, player_id in zip(rows, ids):
        player = by_id.get(player_id)
        if player is None:
            continue
        for field in _PLAYER_EDITABLE_FIELDS:
            if field not in r:
                continue
            value = (r.get(field) or '').strip()
            if field == 'EMA_ID':
                # Same rule as the importer, so an edited id matches an imported one
                # and the EMA export's int() can't meet free text.
                try:
                    value = _normalize_ema_id(value)
                except (TypeError, ValueError):
                    return HttpResponse(
                        f"'{value}' is not a valid EMA number — enter digits only, "
                        f"or leave it blank.", status=400)
            max_length = Player._meta.get_field(field).max_length
            if max_length and len(value) > max_length:
                return HttpResponse(
                    f"{field} is too long: {len(value)} characters "
                    f"(maximum {max_length}).", status=400)
            setattr(player, field, value)
        # First/last are the edited fields; when either changed keep the canonical
        # "First Last" display in sync (bulk_update bypasses the model's save()).
        if 'first_name' in r or 'last_name' in r:
            player.full_name = f"{player.first_name} {player.last_name}".strip()
            names_changed = True
        to_update.append(player)

    if to_update:
        Player.objects.bulk_update(to_update, _PLAYER_EDITABLE_FIELDS + ['full_name'])
        if names_changed:
            # Short names disambiguate across the whole roster ("Chris D." vs
            # "Chris A."), so a name edit can change another competitor's token
            # too: recompute for the tenant, the same way the importer builds them.
            roster = list(Player.objects.filter(tenant=tenant))
            _assign_short_names(roster)
            Player.objects.bulk_update(roster, ['short_name'])
        # Names, countries and teams are all rendered into the cached standings.
        invalidate_leaderboard(tenant.subdomain)
    return HttpResponse('OK')


# Field-size ceiling of the in-app seating generator; also caps one add request.
_MAX_PLAYERS = 200


def _player_row(p):
    """One Player as the editor page's row dict (``_page_player_editor`` and
    ``player_editor_add`` must agree on this shape)."""
    return {'id': p.id, 'draw_number': p.draw_number, 'full_name': p.full_name,
            'first_name': p.first_name, 'last_name': p.last_name,
            'EMA_ID': p.EMA_ID, 'country': p.country, 'team': p.team}


@tenant_admin_required
def player_editor_add(request):
    """Append ``count`` placeholder competitors ("Player 1", "Player 2", …) to the
    player list, for building a tournament in the app instead of importing a
    template. Each placeholder is a mononym (first_name = full_name = short_name),
    not yet drawn in; the organizer renames them in the editor. Numbering
    continues from the current roster size (16 players -> "Player 17"), skipping a
    number only if a competitor already carries that exact name.

    With ``random: true`` in the body the new rows are instead fully made-up
    competitors (name, fake EMA id, country, team when the tournament has teams)
    for rehearsing on a test tournament — refused unless
    ``TournamentSettings.is_test`` is on, so hiding the button is not the only guard."""
    if request.method != 'POST':
        return method_not_allowed()

    tenant = get_tenant(request)
    body = json_body(request)
    tournament = get_tournament(request)
    randomize = bool(body.get('random'))
    if randomize and not tournament.is_test:
        return JsonResponse(
            {'ok': False,
             'error': 'Random players are only available on a test tournament '
                      '(Tournament settings → "This is a test tournament").'},
            status=403)
    count = int_param(body, 'count')
    if not 1 <= count <= _MAX_PLAYERS:
        return JsonResponse(
            {'ok': False, 'error': f'Add between 1 and {_MAX_PLAYERS} players at a time.'},
            status=400)

    existing = Player.objects.filter(tenant=tenant).count()
    if existing + count > _MAX_PLAYERS:
        return JsonResponse(
            {'ok': False,
             'error': f'A tournament holds at most {_MAX_PLAYERS} players '
                      f'({existing} already listed).'},
            status=400)

    taken = set(Player.objects.filter(tenant=tenant).values_list('full_name', flat=True))
    if randomize:
        new_players = _random_players(tenant, count, existing, taken,
                                      with_teams=tournament.has_teams)
        Player.objects.bulk_create(new_players)
        # Real names may share a first name, so re-disambiguate short names across
        # the whole roster (same as the importer and the name editor do).
        roster = list(Player.objects.filter(tenant=tenant).order_by('id'))
        _assign_short_names(roster)
        Player.objects.bulk_update(roster, ['short_name'])
        new_ids = {p.id for p in new_players}
        new_players = [p for p in roster if p.id in new_ids]
    else:
        new_players = []
        k = existing + 1
        while len(new_players) < count:
            name = f'Player {k}'
            if name not in taken:
                new_players.append(Player(
                    tenant=tenant, full_name=name, first_name=name, last_name='',
                    short_name=name, draw_number=None))
            k += 1
        Player.objects.bulk_create(new_players)
    invalidate_leaderboard(tenant.subdomain)
    return JsonResponse({'ok': True, 'players': [_player_row(p) for p in new_players]})


# Name pools for random rehearsal players: deliberately mixed origins so the
# roster looks like an international event, and short so the standings columns
# don't overflow.
_TEST_FIRST_NAMES = (
    'Anna', 'Bo', 'Chen', 'Dara', 'Emil', 'Fatima', 'Greta', 'Hiro', 'Ines', 'Jonas',
    'Kaito', 'Lena', 'Mateo', 'Nadia', 'Oskar', 'Priya', 'Quang', 'Rosa', 'Sven', 'Tomas',
    'Ulla', 'Viktor', 'Wei', 'Xenia', 'Yuki', 'Zara', 'Aiko', 'Bruno', 'Camille', 'Dmitri',
)
_TEST_LAST_NAMES = (
    'Andersson', 'Bauer', 'Costa', 'Dubois', 'Eriksson', 'Fischer', 'Garcia', 'Hansen',
    'Ito', 'Jansen', 'Kowalski', 'Lindqvist', 'Moreau', 'Nakamura', 'Olsen', 'Petrov',
    'Quist', 'Rossi', 'Sato', 'Tanaka', 'Ueda', 'Vogel', 'Wang', 'Xu', 'Yamamoto', 'Zhang',
    'Berg', 'Novak', 'Silva', 'Meyer',
)
_TEST_COUNTRIES = (
    'Sweden', 'France', 'Japan', 'Germany', 'Netherlands', 'Denmark', 'Italy',
    'Poland', 'Austria', 'China',
)


def _random_players(tenant, count, existing, taken, with_teams):
    """Build ``count`` unsaved Player rows with made-up but plausible data.
    Names are unique within the tenant (falls back to a numeric suffix once the
    pools are exhausted — 30×30 combinations comfortably cover _MAX_PLAYERS, but a
    roster can already hold arbitrary names). EMA ids start with ``99`` so they can
    never collide with a real federation number. Teams cycle Team A, B, C… over the
    whole roster so a fresh test tournament ends up with evenly sized teams."""
    rng = random.Random()
    used_ema = set(Player.objects.filter(tenant=tenant).values_list('EMA_ID', flat=True))
    taken = set(taken)
    players = []
    for i in range(count):
        for attempt in range(50):
            first = rng.choice(_TEST_FIRST_NAMES)
            last = rng.choice(_TEST_LAST_NAMES)
            if attempt >= 25:
                last = f'{last} {attempt}'
            full = f'{first} {last}'
            if full not in taken:
                break
        taken.add(full)
        while True:
            ema = f'99{rng.randrange(10**6):06d}'
            if ema not in used_ema:
                break
        used_ema.add(ema)
        team = ''
        if with_teams:
            # 4 players per team, lettered from the roster position.
            team = f'Team {chr(ord("A") + ((existing + i) // 4) % 26)}'
        players.append(Player(
            tenant=tenant, full_name=full, first_name=first, last_name=last,
            short_name=first, EMA_ID=ema, country=rng.choice(_TEST_COUNTRIES),
            team=team, draw_number=None))
    return players


@tenant_admin_required
def player_editor_delete(request):
    """Remove one competitor from the player list. Seats are keyed by draw number,
    not by a FK to the Player, so deleting a drawn-in competitor simply frees the
    slot: the seat shows as "Player #n" again and any score recorded there stays
    with the slot."""
    if request.method != 'POST':
        return method_not_allowed()

    tenant = get_tenant(request)
    player_id = int_param(json_body(request), 'id')
    deleted, _ = Player.objects.filter(tenant=tenant, id=player_id).delete()
    if not deleted:
        return JsonResponse({'ok': False, 'error': 'Unknown competitor'}, status=404)
    invalidate_leaderboard(tenant.subdomain)
    return JsonResponse({'ok': True})


# Public (display screens are public, like /scan and counter_start): serve the
# tenant's uploaded logo. Templates fall back to the static mcr_logo when unset,
# so this is only hit when a logo exists.
def logo(request):
    tournament = get_tournament(request)
    if not tournament.logo:
        raise Http404
    resp = HttpResponse(bytes(tournament.logo), content_type="image/png")
    resp["Cache-Control"] = "public, max-age=86400"
    resp["ETag"] = f'"{tournament.logo_etag}"'
    return resp


@tenant_admin_required
def update_logo(request):
    if request.method != 'POST':
        return method_not_allowed()
    tournament = get_tournament(request)
    if request.POST.get("reset") == "1":
        tournament.logo = None
        tournament.logo_etag = ""
    else:
        f = request.FILES.get("logo")
        if f is None:
            return HttpResponseBadRequest("No file uploaded")
        if f.size > 2 * 1024 * 1024:
            return HttpResponseBadRequest("Logo too large (max 2 MB)")
        data = f.read()
        # Trust the bytes, not the extension: must be a real PNG.
        if data[:8] != b"\x89PNG\r\n\x1a\n":
            return HttpResponseBadRequest("PNG files only")
        tournament.logo = data
        tournament.logo_etag = hashlib.md5(data).hexdigest()
    # Scope the write to the logo fields: a full-row save would also persist this
    # (possibly stale) instance's `counter`, which could stop a running round
    # timer. signals.py still invalidates the cached settings on post_save.
    tournament.save(update_fields=['logo', 'logo_etag'])
    return HttpResponse("OK")


@tenant_admin_and_reauthed
def admin_reset(request):
    """Factory-reset the tournament: wipe every tenant row and its settings.

    Admin-only, POST-only, and re-auth gated (``@tenant_admin_and_reauthed``): the
    operator must have confirmed their password recently, so a stale/borrowed
    session can't drive this irreversible wipe directly. The UI adds a
    type-to-confirm + confirm dialog on top. It clears all tournament data (player list, seating,
    hands, score sheets, published rounds, schedule) *and* the tenant's
    configuration (title/branding/format, logo, screens, screen modes, publish
    target, ceremony state), leaving a blank instance ready for a fresh import.
    Deleting the TournamentSettings row lets get_tournament recreate it at defaults;
    its post_delete signal busts the settings cache and refreshes public displays.
    """
    if request.method != 'POST':
        return method_not_allowed()
    tenant = get_tenant(request)
    if tenant is None:
        return HttpResponseBadRequest("No tournament")
    # The model list lives in tenant_dump (shared with dump/restore). Deleting
    # the settings row resets identity/branding/format, logo and the round timer
    # to defaults; get_tournament recreates a fresh one on next read.
    from ..tenant_dump import wipe_tenant
    with transaction.atomic():
        wipe_tenant(tenant, include_publish_target=True)
    # Wake public displays and scorer pages: nothing is published anymore, the
    # screen set is gone, and the leaderboard/settings caches are stale.
    invalidate_leaderboard(tenant.subdomain)
    broadcast_publish_state(tenant.subdomain, {'published_rounds': []})
    broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screens_changed'})
    return JsonResponse({'status': 'ok'})


@tenant_admin_required
def publish_web(request):
    """Manually regenerate + upload the public static site (the "Publish to web"
    button). Publish also happens automatically on each round publish; this is the
    on-demand re-push. Runs in the background so the request returns at once.

    POST only. CSRF is never checked on a GET, so while this answered one a
    cross-site <img> pointed at it could fire an export and an SFTP upload.
    """
    if request.method != 'POST':
        return method_not_allowed()
    tenant = get_tenant(request)
    subdomain = tenant.subdomain if tenant else ''
    from ..publish.sftp_upload import is_configured
    if not is_configured(subdomain):
        return JsonResponse(
            {'status': 'error',
             'error': 'Static publish is not configured for this tournament '
                      '(add a publish target first).'},
            status=400)
    from ..publish.trigger import fire_static_export
    fire_static_export(subdomain)
    return JsonResponse({'status': 'ok', 'message': 'Publishing to the web…'})


@tenant_role_required('scorer', 'display_op', 'publisher')
def publish_status(request):
    """Poll the running (or last) publish job — drives the shell progress toast.
    Any admin role can read it (auto-publish fires on a publisher's round publish,
    not just staff's manual push). Returns {phase, pct, message, error}; phase is
    idle when nothing is running."""
    tenant = get_tenant(request)
    subdomain = tenant.subdomain if tenant else ''
    from ..publish.trigger import get_progress
    return JsonResponse(get_progress(subdomain) or {'phase': 'idle'})


@tenant_admin_required
def publish_target_save(request):
    """Save this tenant's SFTP publish target (tenant admin only).

    Secrets are write-only: a blank password/key leaves the stored value
    untouched (so the form never has to echo it back), and clear_password /
    clear_key wipe it. Stored encrypted via publish.secrets."""
    if request.method != 'POST':
        return method_not_allowed()
    tenant = get_tenant(request)
    if tenant is None:
        return JsonResponse({'status': 'error', 'error': 'No tournament.'}, status=400)
    from ..models import PublishTarget
    from ..publish import secrets as publish_secrets
    target, _ = PublishTarget.objects.get_or_create(tenant=tenant)

    def _flag(name):
        return request.POST.get(name, '').strip().lower() in ('true', '1', 'on', 'yes')

    target.enabled = _flag('enabled')
    target.host = request.POST.get('host', '').strip()
    target.username = request.POST.get('username', '').strip()
    target.path = request.POST.get('path', '').strip()
    target.backup_path = request.POST.get('backup_path', '').strip()
    target.host_key = request.POST.get('host_key', '').strip()
    port_raw = request.POST.get('port', '').strip() or '22'
    try:
        target.port = int(port_raw)
    except ValueError:
        return JsonResponse(
            {'status': 'error', 'error': f'Port must be a number (got {port_raw!r}).'},
            status=400)

    password = request.POST.get('password', '')
    if request.POST.get('clear_password') == '1':
        target.password_enc = None
    elif password:
        target.password_enc = publish_secrets.encrypt(password)

    key = request.POST.get('private_key', '')
    if request.POST.get('clear_key') == '1':
        target.private_key_enc = None
    elif key.strip():
        target.private_key_enc = publish_secrets.encrypt(key)

    target.save()

    # The advertised spectator URL is a TournamentSettings field (cached and read
    # on every public request), but edited on this page next to the SFTP target.
    tournament = get_tournament(request)
    public_url = request.POST.get('public_url', '').strip()
    if public_url != tournament.public_url:
        tournament.public_url = public_url
        tournament.save(update_fields=['public_url'])  # signals bust the cache
    return JsonResponse({'status': 'ok'})


@tenant_admin_required
def publish_target_test(request):
    """Open + close an SFTP connection using the values currently in the form —
    not the saved target — so staff can verify before saving. A blank password/
    key field falls back to the stored secret, so you can test an unchanged
    credential without re-typing it.

    This does let a tenant admin make the server open a TCP connection to any
    host:port they type. That is inherent to a "test this target" button and the
    role is trusted, so it stays — but each attempt is logged with the user and
    target, so the probing is at least attributable after the fact."""
    if request.method != 'POST':
        return method_not_allowed()
    tenant = get_tenant(request)
    from ..models import PublishTarget
    from ..publish import secrets as publish_secrets
    from ..publish.sftp_upload import PublishConfig, _connect

    host = request.POST.get('host', '').strip()
    if not host:
        return JsonResponse({'status': 'error', 'error': 'Enter a host first.'}, status=400)
    port_raw = request.POST.get('port', '').strip() or '22'
    try:
        port = int(port_raw)
    except ValueError:
        return JsonResponse(
            {'status': 'error', 'error': f'Port must be a number (got {port_raw!r}).'},
            status=400)

    stored = (PublishTarget.objects.filter(tenant=tenant).order_by('id').first()
              if tenant else None)
    password = request.POST.get('password', '')
    if not password and stored and request.POST.get('clear_password') != '1':
        password = publish_secrets.decrypt(stored.password_enc)
    key = request.POST.get('private_key', '')
    if not key.strip() and stored and request.POST.get('clear_key') != '1':
        key = publish_secrets.decrypt(stored.private_key_enc)

    # No `subdomain`: these are unsaved form values, so _connect must not learn
    # and pin a host key from them onto the stored target. PublishConfig.subdomain
    # defaults to '' and _remember_host_key skips on that — keep it that way.
    cfg = PublishConfig(
        host=host, port=port,
        username=request.POST.get('username', '').strip(),
        path=request.POST.get('path', '').strip() or '.',
        password=password,
        key_data=key,
        host_key=request.POST.get('host_key', '').strip(),
    )
    logger.info("publish target test by %s (tenant %s) -> %s:%s",
                request.user, tenant.subdomain if tenant else '-', cfg.host, cfg.port)
    try:
        client = _connect(cfg)
        client.open_sftp().close()
        client.close()
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
    return JsonResponse({'status': 'ok', 'message': f'Connected to {cfg.host}.'})


# The round timer is server-authoritative. `counter` holds an absolute epoch-ms
# "gong moment": the instant the round starts (and the start gong sounds). Before
# it, screens render a synchronized lead window + 3-2-1 countdown; after it, the
# elapsed/remaining time. <= 0 means stopped / never started.
#
# Start = now + LEAD + COUNTDOWN. The LEAD is dead time during which every screen
# has received the broadcast but nothing visible happens yet, so a screen whose
# message is slightly delayed still catches the very first "3" — all rooms count
# down (and gong) in lockstep. Both constants are mirrored as LEAD_MS /
# COUNTDOWN_MS in display_counter.html; keep them in sync.
COUNTER_LEAD_MS = 1000
COUNTER_COUNTDOWN_MS = 3000


def counter_start(request):
    """Read (no args, public — screens poll this) or command the round timer.

    Writes (?action=start|stop) are restricted to display operators and must be
    POSTs: only an explicit admin action may start/stop the counter. Clients never
    supply a timestamp — the server computes it — so projector screens are pure
    renderers and can't race each other or be reset by a stray request.
    """
    tenant = get_tenant(request)
    action = request.GET.get('action')
    if action in ('start', 'stop'):
        if request.method != 'POST' or not has_role(request, 'display_op'):
            return HttpResponseForbidden('forbidden')
        if action == 'start':
            value = int(time.time() * 1000) + COUNTER_LEAD_MS + COUNTER_COUNTDOWN_MS
        else:  # stop / reset
            value = -1
        if tenant is not None:
            set_counter(tenant, value)
            broadcast_display(tenant.subdomain, 'counter.update', {
                'event': 'counter_update',
                'counter': value,
                'server_now': int(time.time() * 1000),
            })
    return JsonResponse({
        'counter': get_counter(tenant),
        'server_now': int(time.time() * 1000),
    })


def _apply_set_tournament(request, tournament, allowed_fields):
    """Persist ``?tournament-<field>=<value>`` params onto the tenant settings and
    return the response to send back. Shared by the Display page (screen-layout
    tuning) and the Tournament settings page (identity + round length), each of
    which passes its own ``allowed_fields`` set — anything not in it is ignored,
    so a display operator can't reach structural or identity fields and no page
    can touch the server-authoritative `counter`."""
    touched_fields = []
    # Fields arrive in the query string historically, and in the POST body from
    # the shared autosave — card CSS runs to kilobytes, which a URL cannot carry
    # reliably. The body wins where both carry a field.
    params = {**request.GET.dict(), **request.POST.dict()}
    for var in params:
        if var.startswith("tournament-"):
            field = var[len("tournament-"):]
            if field not in allowed_fields:
                continue
            if hasattr(tournament, field):
                value = params[var]
                # Coerce booleans: every GET value is a string, and a non-empty
                # string ("false") is truthy — so a raw setattr would store True
                # for both. Map the usual truthy spellings instead.
                if tournament._meta.get_field(field).get_internal_type() == 'BooleanField':
                    value = value.strip().lower() in ('true', '1', 'on', 'yes')
                setattr(tournament, field, value)
                touched_fields.append(field)
    if touched_fields:
        # Reject over-long text here, before it reaches the DB: on
        # PostgreSQL an oversized value raises a bare 500, which the admin
        # UI showed silently. Returning a readable 400 instead lets the
        # page surface exactly which field was too long, and why.
        for field in touched_fields:
            model_field = tournament._meta.get_field(field)
            label = _TOURNAMENT_LABELS.get(field, field)
            max_length = getattr(model_field, "max_length", None)
            value = getattr(tournament, field)
            if max_length and isinstance(value, str) and len(value) > max_length:
                return HttpResponse(
                    f"{label} is too long: {len(value)} characters "
                    f"(maximum {max_length}).",
                    status=400)
            # Field validators (card format/theme/CSS today) — run_validators
            # rather than full_clean or Field.validate: those also enforce
            # blank=False, which would start rejecting the cleared text fields
            # the console has always allowed. to_python first, because every
            # value arrives as a string and a number field's range validators
            # can't compare a string to an int.
            try:
                model_field.run_validators(model_field.to_python(value))
            except ValidationError as exc:
                return HttpResponse(f"{label}: {' '.join(exc.messages)}",
                                    status=400)
        # Scope the write to the fields actually edited: a full-row save would
        # also persist this instance's `counter`, which could stop a running
        # round timer. signals.py still busts the cache on post_save.
        try:
            tournament.save(update_fields=touched_fields)
        except Exception as exc:
            # Any other save failure (e.g. a non-numeric value for a number
            # field) — return the reason rather than a silent 500.
            return HttpResponse(f"Could not save: {exc}", status=400)
    return HttpResponse(str(tournament))


def _save_schedule(request, tenant):
    """Replace the tenant's schedule with the rows posted from the settings editor.

    The editor sends the whole agenda as JSON (ordered), so a full delete+recreate
    keeps row order (``Schedule`` is read ``order_by('id')`` everywhere) without
    tracking per-row ids. Returns the count of round-rows so the UI can flag a
    mismatch with ``nb_rounds`` (the Nth round-row maps to round N)."""
    try:
        rows = json.loads(request.POST.get('schedule', '[]'))
        if not isinstance(rows, list):
            raise ValueError("schedule must be a list")
    except (ValueError, json.JSONDecodeError) as exc:
        return HttpResponse(f"Could not read the schedule: {exc}", status=400)

    max_len = Schedule._meta.get_field('name').max_length
    objs = []
    for row in rows:
        day = (row.get('day') or "").strip()
        time = (row.get('time') or "").strip()
        name = (row.get('name') or "").strip()
        # A wholly blank line is dropped rather than stored as an empty agenda row.
        if not (day or time or name):
            continue
        for value, label in ((day, "Day"), (time, "Time"), (name, "Name")):
            if len(value) > max_len:
                return HttpResponse(
                    f"{label} is too long: {len(value)} characters "
                    f"(maximum {max_len}).", status=400)
        objs.append(Schedule(tenant=tenant, day=day, time=time, name=name,
                             is_round=bool(row.get('is_round'))))

    # One transaction: a failure between the two would leave the tenant with no
    # schedule at all, and since the Nth is_round row *is* round N, an empty agenda
    # silently unaligns every round time in the app rather than raising anywhere.
    with transaction.atomic():
        Schedule.objects.filter(tenant=tenant).delete()
        Schedule.objects.bulk_create(objs)

    # player_rounds (player modal) reads the schedule, and the projector Schedule
    # screen renders it — refresh both.
    invalidate_leaderboard(tenant.subdomain)
    broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screen_update'})

    return JsonResponse({'rounds': sum(1 for o in objs if o.is_round)})


# ---- admin page dispatch --------------------------------------------------
#
# The admin console is one URL (`admin?page=…`): one shell around one page fragment.
# Each page gets a renderer of its own below, and ADMIN_PAGES at the end of the
# section says who may see it. That table is the point of the split — the gates used
# to be a dozen hand-rolled `and not has_role(...)` clauses spread through a single
# 500-line if/elif chain, where a page added without one looked exactly like a page
# that didn't need one. A row without a gate now doesn't parse as a row.
#
# Every renderer takes the same (request, tenant, error) so the dispatcher can call
# them without knowing which is which; only the welcome page has anything to do with
# `error` (the importer redisplays its failure there). A renderer returns either the
# rendered fragment, which the dispatcher wraps in the shell, or an HttpResponse,
# which the dispatcher returns untouched — that is how the two pages that carry
# `?action=` mutations answer with a redirect, a JSON payload or a 405.


def _setup_status(tenant, tournament):
    """The setup checklist's facts, shared by the Setup home, the Print materials
    page and the Run dashboard so they can never disagree about readiness."""
    nb_players = Player.objects.filter(tenant=tenant).count()
    nb_drawn = Player.objects.filter(tenant=tenant, draw_number__isnull=False).count()
    # Whether a seating chart exists at all (imported or generated) — the player
    # list can be drawn in only once there are seats to fill.
    has_seats = Seat.objects.filter(tenant=tenant).exists()
    # The chart's own size (distinct draw slots). Players are added/removed only
    # before play, so a chart of a different size than the list is a setup error
    # the checklist must shout about — the extra people can never be drawn in.
    chart_players = (Seat.objects.filter(tenant=tenant)
                     .values('draw_number').distinct().count()) if has_seats else 0
    # Same story for the chart's round count: nb_rounds can be changed after the
    # chart is built, and a chart of the wrong depth either leaves the last rounds
    # unseated or seats rounds that will never be played.
    chart_rounds = (Seat.objects.filter(tenant=tenant)
                    .values('round_nb').distinct().count()) if has_seats else 0
    # Warn when a schedule exists but its playing rounds don't line up with
    # nb_rounds: the Nth round-row maps to round N (scoring.player_rounds), so a
    # mismatch leaves per-round times blank/misaligned. Only flag once a
    # schedule has been set up — a fresh, empty schedule isn't "wrong".
    schedule_total = Schedule.objects.filter(tenant=tenant).count()
    schedule_rounds = Schedule.objects.filter(tenant=tenant, is_round=True).count()
    chart_mismatch = has_seats and chart_players != nb_players
    chart_round_mismatch = has_seats and chart_rounds != (tournament.nb_rounds or 0)
    # A player is "drawn in" once assigned a draw number; the player list is
    # ready to play when every player holds one.
    draw_done = nb_players > 0 and nb_drawn == nb_players
    return {
        "nb_players": nb_players,
        "nb_drawn": nb_drawn,
        "has_seats": has_seats,
        "chart_players": chart_players,
        "chart_mismatch": chart_mismatch,
        "chart_rounds": chart_rounds,
        "chart_round_mismatch": chart_round_mismatch,
        "draw_done": draw_done,
        "nb_screens": Screen.objects.filter(tenant=tenant).count(),
        "schedule_rounds": schedule_rounds,
        "schedule_round_mismatch": schedule_total > 0 and schedule_rounds != tournament.nb_rounds,
        # The three required steps done (screens are optional): the Setup home
        # offers the way over to Run once this holds.
        "setup_complete": (nb_players > 0 and has_seats and not chart_mismatch
                           and not chart_round_mismatch and draw_done),
    }


def _page_setup_home(request, tenant, error=None):
    """Setup workspace home: the readiness checklist. Also where a failed template
    import reports its error, since importing is a setup action."""
    tournament = get_tournament(request)
    context = {"error": error, "tournament": tournament}
    context.update(_setup_status(tenant, tournament))
    return loader.get_template('mahj/admin_setup_home.html').render(context, request)


def _page_print_materials(request, tenant, error=None):
    tournament = get_tournament(request)
    status = _setup_status(tenant, tournament)
    return loader.get_template('mahj/admin_print_materials.html').render(
        {"tournament": tournament, "uses_teams": tournament.has_teams,
         "draw_done": status["draw_done"],
         "nb_drawn": status["nb_drawn"], "nb_players": status["nb_players"]},
        request,
    )


def _page_card_design(request, tenant, error=None):
    """The printed-player-card design page: format, theme, colours and custom CSS.

    Read-only: every control autosaves through the settings page's
    set_tournament action (the fields are on its allowlist), so there is one
    writer for tournament settings rather than two. The live preview is an iframe
    on the real print page, reloaded after each save — what you see is what will
    print, not a mock-up of it.
    """
    from ..card_themes import CARD_PALETTES, CARD_THEME_DEFAULT_CSS, effective_card_css
    from ..models import CARD_THEMES
    from .print_views import CARD_LAYOUTS

    tournament = get_tournament(request)
    formats = [
        {"key": key,
         "label": layout["label"],
         "per_sheet": layout["per_sheet_label"],
         "cols": layout["cols"],
         "rows": layout["rows"],
         # Card size in mm, so the preview frame matches the printed card.
         "card_w": layout["card_w"],
         "card_h": layout["card_h"]}
        for key, layout in CARD_LAYOUTS.items()
    ]
    return loader.get_template('mahj/admin_card_design.html').render({
        "tournament": tournament,
        "formats": formats,
        # In CARD_THEMES order (classic first, the default), not alphabetical.
        "themes": [t for t in CARD_THEMES if t in CARD_THEME_DEFAULT_CSS],
        "theme_defaults": CARD_THEME_DEFAULT_CSS,
        "palettes": CARD_PALETTES,
        "effective_css": effective_card_css(tournament),
        "card_css_max": CARD_CSS_MAX,
        # Measured against the rendered card: eight rounds reach the bottom of an
        # A7, and nine overflow it. Past that the page warns rather than letting
        # an operator discover the clipping on paper.
        "a7_max_rounds": 8,
    }, request)


def _page_welcome(request, tenant, error=None):
    """Run workspace home: live progress — rounds scored, publish state, the
    round timer — and the manual publish-to-web trigger."""
    from ..publish.sftp_upload import is_configured as _static_publish_configured
    from ..scoring import _last_complete_round, publish_state
    tournament = get_tournament(request)
    # _last_complete_round returns nb_rounds when it finds no incomplete seat —
    # which also happens with no seats at all, a false "all complete". Guard on
    # the seating chart actually existing.
    has_seats = Seat.objects.filter(tenant=tenant).exists()
    complete_round = _last_complete_round(tenant, tournament) if has_seats else 0
    last_published, _ = publish_state(tenant, tournament)
    template2 = loader.get_template('mahj/admin_welcome.html')
    return template2.render(
        {
            "static_publish_enabled": _static_publish_configured(tenant.subdomain if tenant else ''),
            "tournament": tournament,
            "complete_round": complete_round,
            "last_published": last_published,
            # Server-authoritative round timer: >0 and in the future means a
            # round is counting down / running (the dashboard shows it live).
            "counter": get_counter(tenant),
        },
        request,
    )


def _display_action(request, tenant, tournament):
    """Handle a display-page `?action=` mutation. An HttpResponse, or None to render.

    Every action here mutates state (or fires a broadcast), so each must be POST: a
    GET link would let a crafted cross-site navigation drive these — delete a screen,
    blank every projector, rewrite settings — while a display operator is logged in.
    Requiring POST also hands CSRF protection back to Django's middleware, which
    never checks GET.
    """
    action = request.GET.get('action')
    # Every branch below mutates state (or fires a broadcast), so it must be
    # POST: a GET link would let a crafted cross-site navigation drive these
    # (delete a screen, blank every projector, rewrite settings) while a
    # display operator is logged in. Requiring POST also hands CSRF
    # protection back to Django's middleware, which never checks GET.
    is_mutation = (
        action in {"set_tournament", "add_screen", "remove_screen",
                   "identify_screens", "add_mode", "set_all_views"}
        or request.GET.get('rm_mode') or request.GET.get('set_mode'))
    if is_mutation and request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    if action == "set_tournament":
        return _apply_set_tournament(request, tournament, DISPLAY_SETTINGS_FIELDS)
    elif action == "add_screen":
        Screen(tenant=tenant, name="", view="black").save()
        # 'screens_changed' (not plain 'screen_update') so the overview grid
        # redraws for the new screen count. Existing per-screen displays are
        # unaffected: only the last position is ever added or removed.
        broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screens_changed'})
        return HttpResponseRedirect(_display_redirect('screens'))
    elif action == "remove_screen":
        last = Screen.objects.filter(tenant=tenant).order_by('id').last()
        if last is not None:
            last.delete()
            broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screens_changed'})
        return HttpResponseRedirect(_display_redirect('screens'))
    elif action == "identify_screens":
        # Flash each screen's positional number (/1, /2, …) as a corner badge
        # for a few seconds so an operator can match physical projectors to
        # their URLs. Reuses the existing 'screen.update' channel with a
        # distinct event the display socket intercepts without reloading —
        # see mahj/static/js/display_socket.js.
        broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screen_identify'})
        return HttpResponse("")
    elif action == "add_mode":
        mode_name = request.POST.get('mode_name')
        screens = Screen.objects.filter(tenant=tenant).order_by('id')
        views_list = [str(screen.view) for screen in screens]
        ScreenMode(tenant=tenant, name=mode_name, views=views_list).save()
        return HttpResponseRedirect(_display_redirect('screens'))
    elif request.GET.get('rm_mode'):
        mode = _screen_mode_or_404(tenant, request.GET.get('rm_mode'))
        mode.delete()
        return HttpResponseRedirect('admin?page=display')
    elif action == "set_all_views":
        # Bulk "All screens" control: point every screen at one view in a
        # single write + one broadcast (rather than N per-screen posts).
        # Mirrors set_mode's shape so the admin page can sync each card's
        # selects/previews from the returned list.
        view = request.GET.get('view') or 'black'
        screens = Screen.objects.filter(tenant=tenant).order_by('id')
        applied = []
        for screen in screens:
            screen.view = view
            screen.save()
            applied.append({'id': screen.id, 'view': screen.view})
        broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screen_update'})
        return JsonResponse({'screens': applied})
    elif request.GET.get('set_mode'):
        mode = _screen_mode_or_404(tenant, request.GET.get('set_mode'))
        views_list = mode.views if isinstance(mode.views, list) else []
        screens = Screen.objects.filter(tenant=tenant).order_by('id')
        applied = []
        # A mode is a full-room snapshot (add_mode saves every screen's
        # view), so applying one sets every screen: a screen added after
        # the mode was saved goes blank rather than keeping stale content
        # (e.g. live standings during a "Break" mode).
        for i, screen in enumerate(screens):
            screen.view = views_list[i] if i < len(views_list) else "black"
            screen.save()
            applied.append({'id': screen.id, 'view': screen.view})
        broadcast_display(tenant.subdomain, 'screen.update', {'event': 'screen_update'})
        # The admin page applies modes via AJAX so it can refresh the
        # selects/previews in place; a direct (non-AJAX) hit redirects back.
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'screens': applied})
        return HttpResponseRedirect('admin?page=display')
    return None


def _page_display(request, tenant, error=None):
    tournament = get_tournament(request)
    mutated = _display_action(request, tenant, tournament)
    if mutated is not None:
        return mutated
    screens = Screen.objects.filter(tenant=tenant).order_by('id')
    modes = ScreenMode.objects.filter(tenant=tenant).order_by('id')
    # A running ceremony takes over every screen, so the display controls
    # below are inert until it ends. Flag it so the page can warn up front;
    # the banner then tracks live 'ceremony.update' broadcasts over the socket.
    ceremony_state = CeremonyState.objects.filter(tenant=tenant).first()
    context = {
        "screens": screens,
        "modes": _mode_breakdowns(modes, screens),
        "nb_players": Player.objects.filter(tenant=tenant).count(),
        # Distinct non-empty team names — drives the "Standings — teams" page
        # picker (individual pages + team pages).
        "nb_teams": Player.objects.filter(tenant=tenant).exclude(team="")
            .values_list('team', flat=True).distinct().count(),
        "tournament": tournament,
        "ceremony_active": bool(ceremony_state and ceremony_state.phase != 'idle'),
        # Base URL the screens are reachable at, taken from THIS request so the
        # preview iframes stay same-origin (X-Frame-Options: SAMEORIGIN). In the
        # cloud this resolves to https://<tenant>.<BASE_DOMAIN> (proxy headers); in
        # the standalone build to http://<host>:<port>. A hardcoded
        # <BASE_DOMAIN> URL would be cross-origin — and unreachable — on a laptop.
        "screen_base": request.build_absolute_uri('/').rstrip('/'),
        # Which collapsible panel to render already open (see _display_redirect).
        "open_panel": request.GET.get('open', ''),
    }
    # Standalone (LAN-served) build: list the addresses other devices can use
    # to open a screen — loopback (this machine) + the LAN IP. The public IP is
    # filled in client-side (external lookup) with a port-forward warning.
    if settings.STANDALONE:
        from .helpers import lan_ip
        port = request.get_port()
        bases = [("This machine", f"http://127.0.0.1:{port}")]
        ip = lan_ip()
        if ip:
            bases.append(("Same network (LAN)", f"http://{ip}:{port}"))
        context["standalone"] = True
        context["access_bases"] = bases
        context["access_port"] = port
    template2 = loader.get_template('mahj/admin_display.html')
    return template2.render(context, request)


def _page_settings(request, tenant, error=None):
    tournament = get_tournament(request)
    action = request.GET.get('action')
    if action in ("set_tournament", "save_schedule") and request.method != 'POST':
        # Both write; see _display_action on why that means POST.
        return HttpResponseNotAllowed(['POST'])
    if action == "set_tournament":
        return _apply_set_tournament(request, tournament, TOURNAMENT_SETTINGS_FIELDS)
    if action == "save_schedule":
        return _save_schedule(request, tenant)
    template2 = loader.get_template('mahj/admin_settings.html')
    schedule_rows = [
        {"day": s.day or "", "time": s.time or "",
         "name": s.name or "", "is_round": s.is_round}
        for s in Schedule.objects.filter(tenant=tenant).order_by('id')
    ]
    return template2.render({
        "tournament": tournament,
        "schedule_rows": schedule_rows,
    }, request)



def _page_player_editor(request, tenant, error=None):
    players = Player.objects.filter(tenant=tenant).order_by(
        F('draw_number').asc(nulls_last=True), 'full_name')
    player_rows = [_player_row(p) for p in players]
    # The seating-chart slots a draw number may be assigned to (the editor
    # rejects anything else before it reaches admin_player_draw_assign).
    valid_draw_numbers = sorted(set(
        Seat.objects.filter(tenant=tenant).values_list('draw_number', flat=True)))
    template2 = loader.get_template('mahj/admin_player_editor.html')
    return template2.render(
        {"player_rows": player_rows,
         "valid_draw_numbers": valid_draw_numbers,
         # Shows the "Add random players" button; the endpoint re-checks the flag.
         "is_test": get_tournament(request).is_test}, request)



def _page_publish_target(request, tenant, error=None):
    from ..models import PublishTarget
    target = PublishTarget.objects.filter(tenant=tenant).order_by('id').first()
    template2 = loader.get_template('mahj/admin_publish_target.html')
    return template2.render({
        "target": target,
        # Secrets are write-only — never render the value, just whether one
        # is set, so the form can show a "configured" hint.
        "has_password": bool(target and target.password_enc),
        "has_key": bool(target and target.private_key_enc),
        # The advertised spectator URL lives on TournamentSettings (cached,
        # non-secret) but is edited here, next to the SFTP target.
        "public_url": get_tournament(request).public_url,
        "subdomain": tenant.subdomain if tenant else '',
    }, request)



def _page_publisher_overview(request, tenant, error=None):
    tournament = get_tournament(request)
    template2 = loader.get_template('mahj/admin_publisher_overview.html')
    return template2.render({
        "rows": publisher_overview_rows(tenant, tournament),
        "tournament": tournament,
        "subdomain": tenant.subdomain if tenant else '',
    }, request)



def _page_users(request, tenant, error=None):
    # Only this tenant's memberships — other tenants' users are invisible.
    memberships = (Membership.objects.filter(tenant=tenant)
                   .select_related('user').order_by('user__username'))
    # A user is "shared" (credential-managed only by a superuser) if they
    # also belong to another tenant. One query, not one per row.
    shared_user_ids = set(
        Membership.objects.exclude(tenant=tenant)
        .filter(user__in=[m.user_id for m in memberships])
        .values_list('user_id', flat=True))
    user_rows = []
    for m in memberships:
        u = m.user
        user_rows.append({
            "id": u.id,
            "username": u.username,
            "is_tenant_admin": m.is_tenant_admin,
            "is_self": u.id == request.user.id,
            "last_login": u.last_login,
            "has_password": u.has_usable_password(),
            "shared": u.id in shared_user_ids,
            "roles": [{"name": n, "label": TENANT_ROLE_LABELS[n],
                       "active": getattr(m, f'is_{n}')} for n in TENANT_ROLES],
        })
    template2 = loader.get_template('mahj/admin_users.html')
    return template2.render({
        "user_rows": user_rows,
        "role_defs": [{"name": n, "label": TENANT_ROLE_LABELS[n]} for n in TENANT_ROLES],
        # Gates the "add an existing account" form: making an account shared
        # between tournaments is a superuser action (see user_add_existing), and
        # there is nothing to share between in the single-tenant standalone build.
        "can_add_existing": _gate_tenant_management(request),
        "link_validity_days": settings.SESAME_MAX_AGE // 86400,
    }, request)



def _page_tenants(request, tenant, error=None):
    tenant_rows = [
        {"id": t.id, "name": t.name, "subdomain": t.subdomain,
         "admins": Membership.objects.filter(tenant=t, is_tenant_admin=True).count(),
         "members": Membership.objects.filter(tenant=t).count(),
         # Neither of these can be deleted — the fallback every tenant FK
         # points at, and the one this request is being served from. The
         # view refuses both; these just grey the button out rather than
         # offering something that will be rejected.
         "is_default": t.subdomain == Tenant.DEFAULT_SUBDOMAIN,
         "is_current": tenant is not None and t.pk == tenant.pk}
        for t in Tenant.objects.all().order_by('subdomain')
    ]
    template2 = loader.get_template('mahj/admin_tenants.html')
    return template2.render({
        "tenant_rows": tenant_rows,
        "base_domain": settings.BASE_DOMAIN,
    }, request)



def _page_import_template(request, tenant, error=None):
    # The upload confirm dialog names what it will replace, so tell the fragment
    # how big the current tournament is and whether any scores exist. The same
    # flag offers the export's "include scores" checkbox, which has nothing to
    # write until play starts.
    existing_players = Player.objects.filter(tenant=tenant).count()
    existing_scores = scoring_has_started(tenant)
    template2 = loader.get_template('mahj/admin_import_template.html')
    return template2.render({
        'existing_players': existing_players,
        'existing_scores': existing_scores,
    }, request)



def _page_backup(request, tenant, error=None):
    """Backup & restore: download this tournament as a dump file, or replace it
    with an uploaded one (endpoints in views/backup_admin). The restore confirm
    dialog names what it is about to erase, so pass the current size."""
    from ..publish.sftp_upload import is_configured as _static_publish_configured
    template2 = loader.get_template('mahj/admin_backup.html')
    return template2.render({
        'subdomain': tenant.subdomain if tenant else '',
        'existing_players': Player.objects.filter(tenant=tenant).count(),
        'existing_scores': (
            Seat.objects.filter(tenant=tenant, minipoints__isnull=False).exists()
            or Hand.objects.filter(tenant=tenant).exists()
        ),
        # When web publishing is configured, every publish also uploads a dump
        # (outside the served tree) — the page says so.
        'publish_configured': _static_publish_configured(tenant.subdomain if tenant else ''),
    }, request)


def _page_seating(request, tenant, error=None):
    from .. import seating as _seating
    tournament = get_tournament(request)
    nb_players = Player.objects.filter(tenant=tenant).count()
    nb_rounds = tournament.nb_rounds or 0
    # Measure the seating chart currently in place (independent of the player list:
    # its size is the draw slots it seats), so the page can show what exists.
    seats = list(Seat.objects.filter(tenant=tenant)
                 .values_list('round_nb', 'table_nb', 'wind', 'draw_number'))
    current = current_headline = None
    # Draw slots the chart seats — its "number of players". Kept beside the roster
    # size so the page can flag a chart built for a different field.
    n_chart = r_chart = 0
    if seats:
        n_chart = len({s[3] for s in seats})
        r_chart = len({s[0] for s in seats})
        try:
            current = _seating.measure(seats, n_chart, r_chart,
                                       has_teams=tournament.has_teams)
            current_headline = _seating.headline(current)
            current['headline'] = current_headline
        except Exception:
            # measure() derives its key sets from the rows, so it handles a
            # hand-edited or imported chart; anything still raising here is a bug
            # rather than odd data. Keep the page up — a mid-tournament operator
            # needs the rest of it more than the quality panel — but log it, so
            # the panel can't go missing silently the way it used to.
            logger.exception("seating measure() failed for tenant %s",
                             tenant.subdomain if tenant else '-')
            current = None
    can_generate = (nb_players >= 8 and nb_players % 4 == 0 and nb_rounds >= 1)
    seating_scores = (
        Seat.objects.filter(tenant=tenant, minipoints__isnull=False).exists()
        or Hand.objects.filter(tenant=tenant).exists()
    )
    template2 = loader.get_template('mahj/admin_seating.html')
    return template2.render({
        'nb_players': nb_players,
        'nb_rounds': nb_rounds,
        'has_teams': tournament.has_teams,
        'has_seats': bool(seats),
        'chart_players': n_chart,
        'chart_mismatch': bool(seats) and n_chart != nb_players,
        'chart_rounds': r_chart,
        'chart_round_mismatch': bool(seats) and r_chart != nb_rounds,
        'existing_scores': seating_scores,
        'current': current,
        'current_headline': current_headline,
        'can_generate': can_generate,
        'algebraic_ok': can_generate and _seating.algebraic_feasible(nb_players, nb_rounds),
    }, request)



def _page_scoring(request, tenant, error=None):
    tournament = get_tournament(request)
    grid = scores_per_table_grid(request)
    all_players = Player.objects.filter(tenant=tenant).order_by('full_name')
    nb_rounds = rounds_played(scores_per_player_rows(request, True))
    template2 = loader.get_template('mahj/admin_scores_per_table.html')
    published_rounds = list(
        PublishedRound.objects.filter(tenant=tenant)
            .order_by('round_nb').values_list('round_nb', flat=True)
    )
    # The score grid keys its badges by the "<round>-<table>" string its template
    # builds, so ask for that shape.
    validated_keys, filled_keys = _sheet_state_keys(tenant, as_strings=True)
    context = {
        'grid': grid,
        "players": all_players,
        "tournament": tournament,
        # Fed to connectScorerSocket for the live row sync. The helper silently
        # skips the socket when this is empty, so leaving it out breaks sync with
        # no error anywhere — every window shows only its own edits.
        "subdomain": tenant.subdomain if tenant else '',
        # The round the scorer is most likely to want open: the first unscored one.
        # Clamped to a round that actually has a tab — once the last round is fully
        # scored this was nb_rounds + 1, which matched no tab, and since the panes are
        # `x-show="activeRound === N"` with no fallback the whole grid rendered empty.
        # That is exactly when the scorer is reconciling before the ceremony.
        "active_round": min(nb_rounds + 1, max(1, tournament.nb_rounds)),
        "published_rounds": published_rounds,
        "validated_keys": validated_keys,
        "filled_keys": filled_keys,
        # Only publishers (and tenant admins) may publish/unpublish — the
        # endpoint is gated the same way, so keep the toggle disabled for plain
        # scorer accounts to avoid a dead control.
        "can_publish": has_role(request, 'publisher'),
        # The same reasoning for the other direction: this is the publisher's
        # landing page and they may see the grid, but every score mutation is
        # scorer-only. Without this a publisher could type into a cell and watch
        # the save die as an unexplained red pip.
        "user_is_scorer": has_role(request, 'scorer'),
    }
    return template2.render(context, request)



def _page_ceremony(request, tenant, error=None):
    template2 = loader.get_template('mahj/admin_ceremony.html')
    return template2.render({
        "tournament": get_tournament(request),
        "subdomain": tenant.subdomain if tenant else '',
        "screens": Screen.objects.filter(tenant=tenant).order_by('id'),
        # Same-origin base for the preview iframes (see the display page):
        # cloud → https://<tenant>.<BASE_DOMAIN>, standalone → http://<host>:<port>.
        "screen_base": request.build_absolute_uri('/').rstrip('/'),
    }, request)


# ---- who may see each page ------------------------------------------------
#
# The view's own decorator admits any app role, which is the right gate for the
# shell but too wide for most pages inside it: a display operator has no business
# on the score grid, and a scorer none in user management. So each page names its
# own audience.

def _gate_shell(request):
    """Any account the shell itself admits — the decorator is the whole gate."""
    return True


def _gate_tenant_admin(request):
    return is_tenant_admin(request)


def _gate_display_op(request):
    """Display operators only. A scorer or publisher turned away here can't drive
    the page's mutating actions either — they run behind the same gate, with one
    exception: ceremony_control's `action=publish` branch takes the *publisher*
    role instead, because revealing results is a publish. A pure publisher is
    therefore allowed to POST the reveal but is not admitted to this page or to
    ceremony_data, so in practice the ceremony is run by someone holding both
    roles (a tenant admin). The split is an authorization boundary, not a
    usable role separation."""
    return has_role(request, 'display_op')


def _gate_scoring(request):
    """Scorers and publishers. Explicitly not display operators, whom the shell
    admits and who have no reason to see the score grid."""
    return has_role(request, 'scorer', 'publisher')


def _gate_publisher(request):
    """has_role('publisher') also covers tenant admins and superusers."""
    return has_role(request, 'publisher')


def _gate_tenant_management(request):
    """Superuser (not while previewing as a role), and meaningless in the
    single-tenant standalone build (the tenant is pinned via LOCAL_TENANT), so
    it's hidden there rather than left broken."""
    return acting_superuser(request) and not settings.STANDALONE


# `reauth` marks a page a borrowed or unattended session must re-confirm a password
# to reach: these three hold credentials or can wipe the database, so a valid session
# cookie alone isn't enough. `reauth_next` is where the confirm panel returns to —
# None for the user console, which is where it lands by default anyway.
#
# `area` is the workspace the page belongs to. The console is split in two: **Setup**
# (everything done before play — settings, players, seating, printing, accounts)
# and **Run** (everything done during it — scoring, publishing, screens, ceremony).
# The shell renders one sidebar per area; tenant admins switch between them, while
# scorers, publishers and display operators only ever hold Run pages. The area is
# presentation only — access is decided by `gate` alone.
_AdminPage = namedtuple('_AdminPage', 'gate render reauth reauth_next area',
                        defaults=(False, None, 'run'))

ADMIN_PAGES = {
    # Run
    "welcome":            _AdminPage(_gate_shell, _page_welcome),
    "scoring":            _AdminPage(_gate_scoring, _page_scoring),
    "publisher_overview": _AdminPage(_gate_publisher, _page_publisher_overview),
    "display":            _AdminPage(_gate_display_op, _page_display),
    "ceremony":           _AdminPage(_gate_display_op, _page_ceremony),
    # Setup
    "setup":              _AdminPage(_gate_tenant_admin, _page_setup_home, area='setup'),
    "settings":           _AdminPage(_gate_tenant_admin, _page_settings, area='setup'),
    "import_template":    _AdminPage(_gate_tenant_admin, _page_import_template, area='setup'),
    "player_editor":      _AdminPage(_gate_tenant_admin, _page_player_editor, area='setup'),
    "seating":            _AdminPage(_gate_tenant_admin, _page_seating, area='setup'),
    "print_materials":    _AdminPage(_gate_tenant_admin, _page_print_materials, area='setup'),
    "card_design":        _AdminPage(_gate_tenant_admin, _page_card_design, area='setup'),
    "users":              _AdminPage(_gate_tenant_admin, _page_users, reauth=True, area='setup'),
    "tenants":            _AdminPage(_gate_tenant_management, _page_tenants,
                                     reauth=True, reauth_next='tenants', area='setup'),
    "backup":             _AdminPage(_gate_tenant_admin, _page_backup,
                                     reauth=True, reauth_next='backup', area='setup'),
    "publish_target":     _AdminPage(_gate_tenant_admin, _page_publish_target, area='setup'),
}


def _landing_page(request, tenant, tournament):
    """Which page a request with no `?page=` lands on.

    Tenant admins land in Setup (the checklist) until play has started, then on
    the Run dashboard. A single-role account gets the page it actually works from,
    rather than a summary of things it can't reach. Publishers manage publishing
    from the scoring page.

    The order matters for an account holding more than one role: display_op is
    checked before publisher, so someone doing both lands on the screen controls —
    they're standing at the projectors. Ranked on the roles actually *granted*
    (the membership flags), not has_role: publisher implies scorer there, which
    would pull that display-op-plus-publisher onto the scoring page.
    """
    if is_tenant_admin(request):
        from ..scoring import tournament_in_progress
        return "welcome" if tournament_in_progress(tenant, tournament) else "setup"
    m = current_membership(request)
    if m is None:
        return "welcome"
    if m.is_scorer:
        return "scoring"
    if m.is_display_op:
        return "display"
    if m.is_publisher:
        return "scoring"
    return "welcome"


@tenant_role_required('scorer', 'display_op', 'publisher')
def options(request, error=None):
    """The admin console: one shell, one page fragment, dispatched via ADMIN_PAGES."""
    tenant = get_tenant(request)
    if request.GET.get('logout') == "1":
        # Logout is a state change, so it must be POST (a GET link would let a
        # crafted cross-site navigation log the operator out mid-tournament).
        if request.method != 'POST':
            return HttpResponseNotAllowed(['POST'])
        logout(request)
        return HttpResponseRedirect('admin')
    if 'view_as' in request.GET:
        # "View as": an admin previews the console as one of the single-role
        # accounts they hand out (see helpers.current_membership). A state change,
        # so POST only. Keyed on the *real* membership: anyone else is bounced with
        # nothing written, and 'off' always clears.
        if request.method != 'POST':
            return HttpResponseNotAllowed(['POST'])
        role = request.GET['view_as']
        if role in _TENANT_ROLES and real_is_tenant_admin(request):
            request.session[VIEW_AS_SESSION_KEY] = role
        elif role == 'off':
            request.session.pop(VIEW_AS_SESSION_KEY, None)
        return HttpResponseRedirect('admin')

    from ..scoring import tournament_in_progress
    tournament = get_tournament(request)
    # A failed template import (called with `error`) reports on the Setup home,
    # where the import belongs — whatever the admin's landing page would be.
    page = request.GET.get('page') or ('setup' if error else _landing_page(request, tenant, tournament))
    spec = ADMIN_PAGES.get(page)
    if spec is None or not spec.gate(request):
        # "None" is what the shell renders as an empty panel. An unknown page and a
        # page this account may not see are deliberately indistinguishable: probing
        # `?page=…` shouldn't map out what exists.
        page_content = "None"
    elif spec.reauth and not reauth_ok(request):
        page_content = _reauth_gate(request, spec.reauth_next)
    else:
        page_content = spec.render(request, tenant, error)
        if isinstance(page_content, HttpResponse):
            # A mutating `?action=` answered with a redirect, JSON or a 405 — not
            # shell content.
            return page_content

    from ..publish.sftp_upload import is_configured as _static_publish_configured
    # Which sidebar to draw. Only tenant admins can hold a Setup page (every Setup
    # gate is at least _gate_tenant_admin), so a forbidden or unknown page falls
    # back to Run — the workspace every console account has.
    area = spec.area if spec is not None and spec.gate(request) else 'run'
    context = {
        "username": request.user.username,
        "page": page,
        "page_content": page_content,
        "area": area,
        # The View-as menu's choices (helpers._TENANT_ROLES, labelled).
        "view_as_roles": (('scorer', 'Scorer'), ('publisher', 'Publisher'),
                          ('display_op', 'Display operator')),
        # Active-item wash for the sidebar: sky in Setup, amber in Run (the
        # literal class names are listed in admin.html for Tailwind's scanner).
        "active_cls": 'bg-sky-50 text-sky-700' if area == 'setup' else 'bg-amber-50 text-amber-700',
        # The Setup workspace warns once play has started: edits there now affect
        # a live tournament. Only computed for Setup pages (one query set saved on
        # every scoring-page load).
        "setup_in_progress": area == 'setup' and tournament_in_progress(tenant, tournament),
        # user_is_scorer / user_is_display_op / user_is_publisher / is_tenant_admin
        # come from the role_flags context processor (tenant-scoped).
        "uses_teams": tournament.has_teams,
        # Sidebar "Test" badge, so a rehearsal flag left on for a real event is
        # visible on every admin page.
        "is_test": tournament.is_test,
        # Standalone is single-tenant (pinned via LOCAL_TENANT), so the superuser
        # tenant-management page is meaningless there — hide it.
        "standalone": settings.STANDALONE,
        # Drives the shell-wide publish progress toast (polls publish_status) — only
        # when web publishing is configured, so idle installs don't poll.
        "static_publish_enabled": _static_publish_configured(tenant.subdomain if tenant else ''),
    }
    return HttpResponse(loader.get_template('mahj/admin.html').render(context, request))
